"""
Test Excel Formula References Feature
Tests that combined Excel sheets use formulas instead of copied values
"""

import pandas as pd
import sys
import openpyxl
from pathlib import Path

sys.path.insert(0, '/Users/krishangsharma/Downloads/BNC')

from data_joiner import DataJoiner
from excel_formula_exporter import export_combined_excel_with_formulas

print("="*80)
print("TESTING EXCEL FORMULA REFERENCES")
print("="*80)

def test_column_join():
    """Test column join with formula references"""
    print("\n" + "="*80)
    print("TEST 1: Column Join with Formula References")
    print("="*80)
    
    # Create test dataframes
    df1 = pd.DataFrame({
        'Date': ['2023-01-01', '2023-02-01', '2023-03-01'],
        'Revenue': [100, 150, 200]
    })
    
    df2 = pd.DataFrame({
        'Date': ['2023-01-01', '2023-02-01', '2023-03-01'],
        'Users': [1000, 1200, 1500]
    })
    
    source_dfs = {
        'Sales Data': df1,
        'User Metrics': df2
    }
    
    print(f"\nSource DataFrames:")
    print(f"  'Sales Data': {df1.shape}")
    print(f"  'User Metrics': {df2.shape}")
    
    # Join using DataJoiner
    joiner = DataJoiner()
    joined_df = joiner.join_on_column(source_dfs, join_key='Date', mode='reference')
    source_mapping = joiner.get_source_mapping()
    
    print(f"\nJoined DataFrame: {joined_df.shape}")
    print(f"Source Mapping: {source_mapping}")
    
    # Export to Excel
    output_path = '/tmp/test_column_join_formulas.xlsx'
    export_combined_excel_with_formulas(
        joined_df=joined_df,
        source_mapping=source_mapping,
        source_dataframes=source_dfs,
        output_path=output_path,
        join_type='column_join'
    )
    
    print(f"\n✓ Excel file created: {output_path}")
    
    # Verify the Excel file
    wb = openpyxl.load_workbook(output_path)
    print(f"\nSheets in workbook: {wb.sheetnames}")
    
    # Check source sheets exist
    assert 'Sales Data' in wb.sheetnames or 'Sales_Data' in wb.sheetnames, "Sales Data sheet missing"
    assert 'User Metrics' in wb.sheetnames or 'User_Metrics' in wb.sheetnames, "User Metrics sheet missing"
    assert 'Combined_Data' in wb.sheetnames, "Combined_Data sheet missing"
    
    print("✓ All expected sheets present")
    
    # Check combined sheet has formulas
    combined_ws = wb['Combined_Data']
    
    # Check a few cells for formulas
    revenue_cell = combined_ws['B2']  # First Revenue value
    users_cell = combined_ws['C2']    # First Users value
    
    print(f"\nChecking formulas in Combined_Data:")
    print(f"  B2 (Revenue): {revenue_cell.value}")
    print(f"  C2 (Users): {users_cell.value}")
    
    # Verify they are formulas (should start with '=')
    if isinstance(revenue_cell.value, str) and revenue_cell.value.startswith('='):
        print("  ✓ Revenue cell contains formula")
    else:
        print(f"  ❌ Revenue cell should be formula, got: {revenue_cell.value}")
        return False
    
    if isinstance(users_cell.value, str) and users_cell.value.startswith('='):
        print("  ✓ Users cell contains formula")
    else:
        print(f"  ❌ Users cell should be formula, got: {users_cell.value}")
        return False
    
    print("\n🎉 TEST 1 PASSED: Column join formulas working!")
    return True


def test_row_append():
    """Test row append with formula references"""
    print("\n" + "="*80)
    print("TEST 2: Row Append with Formula References")
    print("="*80)
    
    # Create test dataframes
    df1 = pd.DataFrame({
        'Product': ['A', 'B'],
        'Sales': [100, 150]
    })
    
    df2 = pd.DataFrame({
        'Product': ['C', 'D'],
        'Sales': [200, 250]
    })
    
    source_dfs = {
        'Q1 Data': df1,
        'Q2 Data': df2
    }
    
    print(f"\nSource DataFrames:")
    print(f"  'Q1 Data': {df1.shape}")
    print(f"  'Q2 Data': {df2.shape}")
    
    # Append using DataJoiner
    joiner = DataJoiner()
    appended_df = joiner.append_rows(source_dfs, mode='reference')
    source_mapping = joiner.get_source_mapping()
    
    print(f"\nAppended DataFrame: {appended_df.shape}")
    print(f"Source Mapping (first 2 rows): {dict(list(source_mapping.items())[:2])}")
    
    # Export to Excel
    output_path = '/tmp/test_row_append_formulas.xlsx'
    export_combined_excel_with_formulas(
        joined_df=appended_df,
        source_mapping=source_mapping,
        source_dataframes=source_dfs,
        output_path=output_path,
        join_type='row_append'
    )
    
    print(f"\n✓ Excel file created: {output_path}")
    
    # Verify the Excel file
    wb = openpyxl.load_workbook(output_path)
    print(f"\nSheets in workbook: {wb.sheetnames}")
    
    # Check combined sheet has formulas
    combined_ws = wb['Combined_Data']
    
    # Check cells from first sheet (rows 2-3)
    cell_a2 = combined_ws['A2']  # Product A
    cell_b2 = combined_ws['B2']  # Sales 100
    
    # Check cells from second sheet (rows 4-5)
    cell_a4 = combined_ws['A4']  # Product C
    cell_b4 = combined_ws['B4']  # Sales 200
    
    print(f"\nChecking formulas in Combined_Data:")
    print(f"  A2 (from Q1): {cell_a2.value}")
    print(f"  B2 (from Q1): {cell_b2.value}")
    print(f"  A4 (from Q2): {cell_a4.value}")
    print(f"  B4 (from Q2): {cell_b4.value}")
    
    # Verify they are formulas
    formulas_found = 0
    for cell in [cell_a2, cell_b2, cell_a4, cell_b4]:
        if isinstance(cell.value, str) and cell.value.startswith('='):
            formulas_found += 1
    
    if formulas_found >= 3:  # At least most cells should have formulas
        print(f"  ✓ Found {formulas_found}/4 formula references")
        print("\n🎉 TEST 2 PASSED: Row append formulas working!")
        return True
    else:
        print(f"  ❌ Only {formulas_found}/4 cells have formulas")
        return False


def test_sheet_name_sanitization():
    """Test that sheet names with spaces/special chars are handled"""
    print("\n" + "="*80)
    print("TEST 3: Sheet Name Sanitization")
    print("="*80)
    
    # Create dataframes with problematic names
    df1 = pd.DataFrame({'A': [1, 2], 'B': [3, 4]})
    df2 = pd.DataFrame({'A': [1, 2], 'C': [5, 6]})
    
    source_dfs = {
        'Revenue: 2023 [Final]': df1,  # Has special chars
        'User Data (Q1)': df2           # Has parens
    }
    
    print(f"\nSource DataFrames with special characters:")
    for name in source_dfs.keys():
        print(f"  '{name}'")
    
    # Join
    joiner = DataJoiner()
    joined_df = joiner.join_on_column(source_dfs, join_key='A', mode='reference')
    source_mapping = joiner.get_source_mapping()
    
    # Export
    output_path = '/tmp/test_sheet_names.xlsx'
    export_combined_excel_with_formulas(
        joined_df=joined_df,
        source_mapping=source_mapping,
        source_dataframes=source_dfs,
        output_path=output_path,
        join_type='column_join'
    )
    
    # Verify
    wb = openpyxl.load_workbook(output_path)
    print(f"\nSanitized sheets: {wb.sheetnames}")
    
    # Check that special characters are replaced
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Combined_Data':
            continue
        # Should not contain [ ] : * ? / \
        invalid_chars = ['[', ']', ':', '*', '?', '/', '\\']
        has_invalid = any(char in sheet_name for char in invalid_chars)
        if has_invalid:
            print(f"  ❌ Sheet '{sheet_name}' contains invalid characters")
            return False
    
    print("  ✓ All sheet names sanitized correctly")
    print("\n🎉 TEST 3 PASSED: Sheet name sanitization working!")
    return True


# Run all tests
if __name__ == "__main__":
    results = []
    
    try:
        results.append(("Column Join", test_column_join()))
    except Exception as e:
        print(f"\n❌ TEST 1 FAILED: {e}")
        import traceback
        traceback.print_exc()
        results.append(("Column Join", False))
    
    try:
        results.append(("Row Append", test_row_append()))
    except Exception as e:
        print(f"\n❌ TEST 2 FAILED: {e}")
        import traceback
        traceback.print_exc()
        results.append(("Row Append", False))
    
    try:
        results.append(("Sheet Sanitization", test_sheet_name_sanitization()))
    except Exception as e:
        print(f"\n❌ TEST 3 FAILED: {e}")
        import traceback
        traceback.print_exc()
        results.append(("Sheet Sanitization", False))
    
    # Summary
    print("\n" + "="*80)
    print("TEST SUMMARY")
    print("="*80)
    
    for test_name, passed in results:
        status = "✅ PASS" if passed else "❌ FAIL"
        print(f"{status}: {test_name}")
    
    total_passed = sum(1 for _, passed in results if passed)
    total_tests = len(results)
    
    print(f"\nTotal: {total_passed}/{total_tests} tests passed")
    
    if total_passed == total_tests:
        print("\n🎉 ALL TESTS PASSED - Excel Formula References Working!")
        print("\nGenerated test files:")
        print("  - /tmp/test_column_join_formulas.xlsx")
        print("  - /tmp/test_row_append_formulas.xlsx")
        print("  - /tmp/test_sheet_names.xlsx")
        print("\nOpen these files in Excel to verify:")
        print("  1. Click any cell in Combined_Data sheet")
        print("  2. Check formula bar shows reference like ='Sheet1'!A1")
        print("  3. Ctrl+Click to jump to source")
        print("  4. Change value in source → combined updates")
    else:
        print(f"\n❌ {total_tests - total_passed} TEST(S) FAILED")
        sys.exit(1)
