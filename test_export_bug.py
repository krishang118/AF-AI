#!/usr/bin/env python3
"""
Testing Excel Export Functions
The join works - testing if export functions are the issue
"""

import pandas as pd
import tempfile
from pathlib import Path
from data_joiner import DataJoiner
from excel_formula_exporter import export_combined_excel_with_formulas
from hybrid_excel_export import export_hybrid_excel
from data_editor import DataEditor
import openpyxl

print("="*80)
print("TESTING EXCEL EXPORT FUNCTIONS")
print("="*80)

# Create test data
df1 = pd.DataFrame({
    'period': ['2020', '2021', '2022'],
    'Revenue': [100, 110, 121]
})
df2 = pd.DataFrame({
    'period': ['2020', '2021', '2022'],
    'Cost': [40, 45, 50]
})

print("\nOriginal Data:")
print("DF1:", df1.to_dict())
print("DF2:", df2.to_dict())

# Perform join
joiner = DataJoiner()
combined = joiner.join_on_column({'Sheet1': df1, 'Sheet2': df2}, 'period', mode='reference')

print("\nCombined Data:")
print(combined)
print("Combined values:", combined.values)

# Test 1: Formula export
print("\n\n1. Testing FORMULA EXPORT...")
try:
    output_formula = tempfile.mktemp(suffix='_formula.xlsx')
    
    export_combined_excel_with_formulas(
        joined_df=combined,
        source_mapping=joiner.get_source_mapping(),
        source_dataframes={'Sheet1': df1, 'Sheet2': df2},
        output_path=output_formula,
        join_type='column_join'
    )
    
    print(f"Exported to: {output_formula}")
    
    # Read back and check
    wb = openpyxl.load_workbook(output_formula)
    ws = wb['Combined_Data']
    
    print("\nCombined_Data sheet values:")
    for row_idx in range(1, min(5, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            row_values.append(f"{cell.value} (formula: {cell.value if str(cell.value).startswith('=') else 'NO'})")
        print(f"Row {row_idx}: {row_values}")
    
    wb.close()
    
    # Read with pandas
    df_formula = pd.read_excel(output_formula, sheet_name='Combined_Data')
    print("\nRead back with pandas:")
    print(df_formula)
    
    if (df_formula.select_dtypes(include=['number']) == 0).all().all():
        print("\n🚨 BUG CONFIRMED IN FORMULA EXPORT: ALL VALUES ARE ZERO!")
    else:
        print("\n✅ Formula export values OK")
        
except Exception as e:
    print(f"❌ Error in formula export: {e}")
    import traceback
    traceback.print_exc()

# Test 2: Hybrid export
print("\n\n2. Testing HYBRID EXPORT...")
try:
    output_hybrid = tempfile.mktemp(suffix='_hybrid.xlsx')
    
    # Make a small edit
    editor = DataEditor(combined)
    editor.mark_as_na([1], ['Revenue'])
    affected = editor.get_affected_cells()
    
    export_hybrid_excel(
        combined_df=editor.df,
        affected_cells=affected,
        source_mapping=joiner.get_source_mapping(),
        source_dataframes={'Sheet1': df1, 'Sheet2': df2},
        output_path=output_hybrid,
        join_type='column_join'
    )
    
    print(f"Exported to: {output_hybrid}")
    
    # Read back
    df_hybrid = pd.read_excel(output_hybrid, sheet_name='Combined_Data')
    print("\nRead back with pandas:")
    print(df_hybrid)
    
    if (df_hybrid.select_dtypes(include=['number']) == 0).all().all():
        print("\n🚨 BUG CONFIRMED IN HYBRID EXPORT: ALL VALUES ARE ZERO!")
    else:
        print("\n✅ Hybrid export values OK")
        
except Exception as e:
    print(f"❌ Error in hybrid export: {e}")
    import traceback
    traceback.print_exc()

# Test 3: Simple pandas export (control test)
print("\n\n3. Testing SIMPLE PANDAS EXPORT (control)...")
try:
    output_simple = tempfile.mktemp(suffix='_simple.xlsx')
    combined.to_excel(output_simple, index=False)
    
    df_simple = pd.read_excel(output_simple)
    print("Read back:")
    print(df_simple)
    
    if (df_simple.select_dtypes(include=['number']) == 0).all().all():
        print("\n🚨 BUG IN SIMPLE EXPORT TOO!")
    else:
        print("\n✅ Simple export OK")
        
except Exception as e:
    print(f"❌ Error: {e}")

print("\n" + "="*80)
