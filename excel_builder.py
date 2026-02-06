"""
Excel Reference Builder Module
Generates Excel files with formula references instead of static values
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any
from pathlib import Path


class ExcelReferenceBuilder:
    """Builds Excel files with live formula references to source sheets"""
    
    def __init__(self):
        self.workbook = None
        
    def create_reference_table(self, source_dataframes: Dict[str, pd.DataFrame],
                              combined_df: pd.DataFrame,
                              source_mapping: Dict,
                              output_path: Path,
                              join_mode: str = 'join') -> Path:
        """
        Creates an Excel file with formula references to source sheets
        
        Args:
            source_dataframes: Original source dataframes {sheet_name: df}
            combined_df: The combined/joined dataframe
            source_mapping: Mapping from DataJoiner showing source origins
            output_path: Where to save the Excel file
            join_mode: 'join' (column-based) or 'append' (row-based)
        
        Returns:
            Path to created Excel file
        """
        # Create new workbook with proper initialization
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Add combined sheet FIRST (so it's the default view)
        if join_mode == 'join':
            self._add_joined_sheet_with_formulas(combined_df, source_mapping)
        else:  # append
            self._add_appended_sheet_with_formulas(combined_df, source_mapping)
        
        # Add source sheets after combined sheet
        for sheet_name, df in source_dataframes.items():
            self._add_source_sheet(sheet_name, df)
        
        # Set workbook properties
        self.workbook.properties.title = "Combined Data with References"
        self.workbook.properties.creator = "BNC Data Wizard"
        
        # Save workbook
        self.workbook.save(output_path)
        return output_path
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name to be Excel-compatible"""
        from pathlib import Path
        
        # Remove file extension if present
        name = Path(name).stem
        
        # Excel sheet names: max 31 chars, no: \ / ? * [ ]
        invalid_chars = ['\\', '/', '?', '*', '[', ']']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')
        
        # Truncate to 31 characters
        return sanitized[:31]
    
    def _add_source_sheet(self, sheet_name: str, df: pd.DataFrame):
        """Add a source sheet with data values"""
        safe_name = self._sanitize_sheet_name(f"Src_{sheet_name}")
        ws = self.workbook.create_sheet(safe_name)
        
        # Add headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = str(col_name)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Convert value to native Python type if needed
                if pd.isna(value):
                    cell.value = None
                elif isinstance(value, (pd.Timestamp, pd.DatetimeTZDtype)):
                    cell.value = value.to_pydatetime() if hasattr(value, 'to_pydatetime') else value
                else:
                    cell.value = value
        
        # Adjust column widths safely
        for col_idx in range(1, len(df.columns) + 1):
            try:
                max_length = len(str(df.columns[col_idx - 1]))
                # Check first few data rows for width
                for row_idx in range(2, min(10, len(df) + 2)):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            except:
                pass  # Skip if any issues
    
    def _add_joined_sheet_with_formulas(self, combined_df: pd.DataFrame, 
                                       source_mapping: Dict):
        """Add combined sheet with formulas referencing source sheets (for column joins)"""
        ws = self.workbook.create_sheet("Combined", 0)  # Index 0 = first sheet
        
        # Add headers
        for col_idx, col_name in enumerate(combined_df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = str(col_name)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Build column-to-source mapping
        col_to_source = {}
        for col_name in combined_df.columns:
            source_sheet = source_mapping.get(col_name)
            if source_sheet:
                safe_source = self._sanitize_sheet_name(f"Src_{source_sheet}")
                col_to_source[col_name] = safe_source
        
        # Add HYPERLINK formulas for click-through navigation
        for row_idx in range(len(combined_df)):
            for col_idx, col_name in enumerate(combined_df.columns, start=1):
                cell = ws.cell(row=row_idx + 2, column=col_idx)
                
                source_sheet_name = col_to_source.get(col_name)
                
                if source_sheet_name:
                    # Create HYPERLINK formula for click-through navigation
                    col_letter = get_column_letter(col_idx)
                    excel_row = row_idx + 2
                    
                    # Build the cell reference
                    cell_ref = f"'{source_sheet_name}'!{col_letter}{excel_row}"
                    
                    # Use INDIRECT to get the value, HYPERLINK for navigation
                    # This ensures the cell shows the actual value and is clickable
                    formula = f'=HYPERLINK("#{cell_ref}", INDIRECT("{cell_ref}"))'
                    
                    cell.value = formula
                else:
                    # No source mapping, use actual value
                    value = combined_df.iloc[row_idx, col_idx - 1]
                    if pd.isna(value):
                        cell.value = None
                    else:
                        cell.value = value
        
        # Adjust column widths
        for col_idx in range(1, len(combined_df.columns) + 1):
            try:
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 15
            except:
                pass
    
    def _add_appended_sheet_with_formulas(self, combined_df: pd.DataFrame,
                                         source_mapping: Dict):
        """Add combined sheet with formulas referencing source sheets (for row appends)"""
        ws = self.workbook.create_sheet("Combined", 0)  # Index 0 = first sheet
        
        # Add headers
        for col_idx, col_name in enumerate(combined_df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = str(col_name)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add HYPERLINK formulas for click-through navigation
        for result_row_idx in range(len(combined_df)):
            # Get source mapping for this row
            row_mapping = source_mapping.get(result_row_idx, {})
            source_sheet = row_mapping.get('sheet')
            source_row = row_mapping.get('source_row', 0)
            
            for col_idx, col_name in enumerate(combined_df.columns, start=1):
                cell = ws.cell(row=result_row_idx + 2, column=col_idx)
                
                if source_sheet:
                    safe_source = self._sanitize_sheet_name(f"Src_{source_sheet}")
                    col_letter = get_column_letter(col_idx)
                    excel_row = source_row + 2  # +2 for header and 0-indexing
                    
                    # Build cell reference
                    cell_ref = f"'{safe_source}'!{col_letter}{excel_row}"
                    
                    # HYPERLINK with INDIRECT for display
                    formula = f'=HYPERLINK("#{cell_ref}", INDIRECT("{cell_ref}"))'
                    
                    cell.value = formula
                else:
                    # No source mapping, use actual value
                    value = combined_df.iloc[result_row_idx, col_idx - 1]
                    if pd.isna(value):
                        cell.value = None
                    else:
                        cell.value = value
        
        # Adjust column widths
        for col_idx in range(1, len(combined_df.columns) + 1):
            try:
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 15
            except:
                pass


def create_copy_table(combined_df: pd.DataFrame, output_path: Path, 
                      source_dataframes: Dict[str, pd.DataFrame] = None) -> Path:
    """
    Creates a simple Excel file with copied values (no formulas)
    
    Args:
        combined_df: The combined dataframe
        output_path: Where to save the Excel file
        source_dataframes: Optional dict of source dataframes to include as extra sheets
    
    Returns:
        Path to created Excel file
    """
    # Use pandas to_excel for simple copy mode
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        combined_df.to_excel(writer, sheet_name='Combined', index=False)
        
        # Optional: Add basic styling
        workbook = writer.book
        worksheet = writer.sheets['Combined']
        
        # Style header
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        for col_idx in range(1, len(combined_df.columns) + 1):
            try:
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 15
            except:
                pass
        
        # Add Source Sheets if provided
        if source_dataframes:
            for name, df in source_dataframes.items():
                safe_name = f"Src_{name}"[:31]  # Excel limit
                # Remove invalid chars
                for char in ['\\', '/', '?', '*', '[', ']']:
                    safe_name = safe_name.replace(char, '_')
                
                df.to_excel(writer, sheet_name=safe_name, index=False)
                
                # Style source sheets too
                if safe_name in writer.sheets:
                    src_ws = writer.sheets[safe_name]
                    for cell in src_ws[1]:
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Set workbook properties
        workbook.properties.title = "Combined Data"
        workbook.properties.creator = "BNC Data Wizard"
    
    return output_path
