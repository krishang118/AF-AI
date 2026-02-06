"""
Test to understand how new rows with NA values are exported
"""
import pandas as pd
from hybrid_excel_export import export_hybrid_excel
import openpyxl

# Create source data
source_df = pd.DataFrame({
    'period': ['2022-Q1', '2022-Q2', '2022-Q3'],
    'revenue': [100, 200, 300]
})

# Create combined data with a NEW ROW added (all NA except one cell edited)
combined_df = pd.DataFrame({
    'period': ['2022-Q1', '2022-Q2', '2022-Q3', '2022-Q4'],  # New row added
    'revenue': [100, 200, 300, pd.NA]  # NA value
})

# Simulate: user edited the period cell in the new row
combined_df.loc[3, 'period'] = '2022-Q4'  # This was edited from NA

# Affected cells: only (3, 'period') was edited, (3, 'revenue') should stay NA
affected_cells = {(3, 'period')}

# Source mapping for column join (maps columns to source sheets)
source_mapping = {
    'period': 'Source1',
    'revenue': 'Source1'
}

source_dataframes = {'Source1': source_df}

# Export
output_path = '/tmp/test_new_row_na.xlsx'
export_hybrid_excel(
    combined_df=combined_df,
    affected_cells=affected_cells,
    source_mapping=source_mapping,
    source_dataframes=source_dataframes,
    output_path=output_path,
    join_type='column_join'
)

# Read back and check
wb = openpyxl.load_workbook(output_path)
ws = wb['Combined_Data']

print("Row 5 (new row, index 3):")
print(f"  A5 (period): {ws['A5'].value} (edited)")
print(f"  B5 (revenue): {ws['B5'].value} (should be None/blank, NOT 0)")

# Check if B5 has a formula
if ws['B5'].value is not None:
    print(f"  B5 has value: {ws['B5'].value}")
if hasattr(ws['B5'], 'formula') and ws['B5'].value:
    print(f"  B5 formula: {ws['B5'].value}")

wb.close()
