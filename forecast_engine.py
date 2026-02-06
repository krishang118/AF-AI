"""
Assumption-Driven Forecasting Engine
Core backend for data ingestion, normalization, assumptions, and deterministic forecasting.
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional, Any
from dataclasses import dataclass, asdict
from enum import Enum
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from docx import Document


class AssumptionType(Enum):
    GROWTH = "growth"
    EVENT = "event"
    PRICING = "pricing"
    MARKET_SHARE = "market_share"
    EXTERNAL_SHOCK = "external_shock"
    CAPACITY = "capacity"


class EventType(Enum):
    PRODUCT_LAUNCH = "product_launch"
    PRICE_CHANGE = "price_change"
    MARKET_ENTRY = "market_entry"
    MARKET_EXIT = "market_exit"
    REGULATION = "regulation"
    CAPACITY_EXPANSION = "capacity_expansion"


class DependencyType(Enum):
    """Types of dependencies between assumptions"""
    REQUIRES = "requires"      # Only applies if dependency exists
    MODIFIES = "modifies"      # Scales another assumption
    OFFSETS = "offsets"        # Subtracts from another assumption
    REPLACES = "replaces"      # Overrides another assumption


class AssumptionLayer:
    """Processing order for assumptions (lower = processed first)"""
    STRUCTURAL_EVENTS = 0    # Launches, regulation
    GROWTH_BASELINE = 1      # Core growth assumptions
    MARKET_DYNAMICS = 2      # Pricing, market share
    EXTERNAL_SHOCKS = 3      # Black swans, risks


@dataclass
class Assumption:
    """First-class assumption object (v2 enhanced)"""
    id: str
    type: AssumptionType
    name: str
    metric: str  # Which metric this applies to
    value: float  # Primary value (growth rate, multiplier, etc.) - midpoint if using ranges
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    confidence: str = "medium"  # low, medium, high
    source: str = "analyst"  # analyst, data, llm
    notes: str = ""
    depends_on: Optional[str] = None  # ID of event/assumption this depends on (legacy)
    apply_if: str = "always"  # "always", "if_occurs", "if_not_occurs" (legacy)
    
    # v2 enhancements
    min_value: Optional[float] = None  # Low end of range
    max_value: Optional[float] = None  # High end of range
    use_range: bool = False  # Use ranges instead of confidence multipliers
    layer: int = 1  # Processing order (see AssumptionLayer)
    dependency_type: str = "requires"  # Type from DependencyType enum (stored as string for compatibility)
    dependency_id: Optional[str] = None  # New dependency field (replaces depends_on)
    condition: Optional[Dict[str, Any]] = None  # Conditional logic
    
    def to_dict(self):
        result = {**asdict(self), 'type': self.type.value}
        # Convert dependency_type enum if needed
        if hasattr(self.dependency_type, 'value'):
            result['dependency_type'] = self.dependency_type.value
        return result


@dataclass
class Event:
    """Discrete event that affects forecast"""
    id: str
    event_type: EventType
    name: str
    metric: str
    date: str
    impact_multiplier: float  # e.g., 1.15 = 15% boost
    decay_periods: int = 0  # How long effect lasts (0 = permanent)
    notes: str = ""
    
    def to_dict(self):
        return {**asdict(self), 'event_type': self.event_type.value}


@dataclass
class GrowthRegime:
    """Piecewise growth regime for more realistic forecasting"""
    id: str
    metric: str
    start_date: str
    end_date: Optional[str]  # None = ongoing
    growth_rate: float
    name: str
    regime_type: str = "detected"  # "detected" or "analyst_defined"
    notes: str = ""
    
    def to_dict(self):
        return asdict(self)


class DataNormalizer:
    """Handles file ingestion and normalization"""
    
    COMMON_DATE_COLS = ['date', 'period', 'year', 'yr', 'fy', 'month', 'quarter', 'time', 'timestamp', 'epoch', 'step', 'week', 'day']
    METRIC_SYNONYMS = {
        'revenue': ['revenue', 'sales', 'net_sales', 'total_sales', 'turnover'],
        'users': ['users', 'customers', 'active_users', 'subscribers'],
        'volume': ['volume', 'units', 'quantity', 'qty'],
        'price': ['price', 'unit_price', 'avg_price', 'pricing'],
        'market_share': ['market_share', 'share', 'market_position'],
    }
    
    def __init__(self):
        self.data_frames: Dict[str, pd.DataFrame] = {}
        self.metadata: Dict[str, Dict] = {}
        self.master_df: Optional[pd.DataFrame] = None
        
    def load_files(self, file_paths: List[Path]) -> Dict[str, List[str]]:
        """Load all uploaded files and return warnings"""
        warnings = {"errors": [], "warnings": [], "info": []}
        
        for path in file_paths:
            try:
                if path.suffix.lower() == '.csv':
                    df = pd.read_csv(path)
                    self.data_frames[path.stem] = df
                    warnings["info"].append(f"Loaded CSV: {path.name} ({len(df)} rows)")
                    
                elif path.suffix.lower() in ['.xlsx', '.xls']:
                    xl_file = pd.ExcelFile(path)
                    num_sheets = len(xl_file.sheet_names)
                    
                    for sheet_name in xl_file.sheet_names:
                        df = pd.read_excel(path, sheet_name=sheet_name)
                        
                        # Better naming: Use filename for single-sheet Excel (like CSV)
                        # For multi-sheet, use filename_sheetname to avoid conflicts
                        if num_sheets == 1:
                            key = path.stem  # Use filename for single-sheet Excel
                        else:
                            key = f"{path.stem}_{sheet_name}"  # filename_sheetname for multi-sheet
                        
                        self.data_frames[key] = df
                        warnings["info"].append(f"Loaded '{key}' from {path.name} ({len(df)} rows)")
                        
                elif path.suffix.lower() == '.docx':
                    text = self._extract_docx_text(path)
                    self.metadata[path.stem] = {'type': 'document', 'content': text}
                    warnings["info"].append(f"Loaded document: {path.name}")
                    
            except Exception as e:
                warnings["errors"].append(f"Failed to load {path.name}: {str(e)}")
        
        return warnings
    
    def _extract_docx_text(self, path: Path) -> str:
        """Extract text from Word document"""
        doc = Document(path)
        return "\n".join([para.text for para in doc.paragraphs])
    
    def normalize_columns(self) -> Dict[str, List[str]]:
        """Normalize column names across all dataframes"""
        warnings = {"renamed": [], "standardized": []}
        
        for name, df in self.data_frames.items():
            original_cols = df.columns.tolist()
            
            # Clean column names
            df.columns = df.columns.str.lower().str.strip()
            df.columns = df.columns.str.replace(r'[^\w\s]', '', regex=True)
            df.columns = df.columns.str.replace(r'\s+', '_', regex=True)
            
            # Map to standard names with deduplication
            new_cols = []
            seen_cols = {}
            
            for col in df.columns:
                mapped = False
                final_name = col
                
                # Try synonyms
                for standard, synonyms in self.METRIC_SYNONYMS.items():
                    if any(syn in col for syn in synonyms):
                        final_name = standard
                        mapped = True
                        warnings["standardized"].append(f"{name}: {col} -> {standard}")
                        break
                
                if not mapped:
                    final_name = col
                
                # Handle duplicates by appending _2, _3, etc.
                if final_name in seen_cols:
                    seen_cols[final_name] += 1
                    final_name = f"{final_name}_{seen_cols[final_name]}"
                    warnings["renamed"].append(f"{name}: Duplicate {col} renamed to {final_name}")
                else:
                    seen_cols[final_name] = 1
                
                new_cols.append(final_name)
            
            df.columns = new_cols
            self.data_frames[name] = df
        
        return warnings
    
    def detect_date_column(self, df: pd.DataFrame) -> Optional[str]:
        """Detect which column represents dates/periods"""
        # 1. Check heuristics (Name)
        for col in df.columns:
            col_lower = str(col).lower()
            if any(date_word in col_lower for date_word in self.COMMON_DATE_COLS):
                return col
        
        # 2. Check content (Values)
        for col in df.columns:
            try:
                # Check for numeric years (1900-2100)
                is_numeric = pd.api.types.is_numeric_dtype(df[col])
                if is_numeric:
                    min_val = df[col].min()
                    max_val = df[col].max()
                    # Year Range
                    if 1900 <= min_val <= 2100: 
                        return col
                    # Timestamp Range (e.g. seconds since epoch > 1980)
                    if min_val > 300000000: 
                         return col
                    continue
                
                # Check for String patterns (FY23, 2023E, etc.)
                if pd.api.types.is_string_dtype(df[col]):
                    sample = df[col].dropna().head(10).astype(str)
                    if len(sample) == 0: continue
                    
                    # Regex for "FY" or "Year" style
                    # e.g. FY2020, 2020E, 2020A, Y2020
                    has_years = sample.str.contains(r'(?:19|20)\d{2}', regex=True).all()
                    if has_years:
                        return col
                        
                # If Strict String Date, try parsing
                pd.to_datetime(df[col].head(), errors='coerce')
                if df[col].head().notna().sum() > 0:
                    return col
            except:
                continue
        return None
    
    def unify_data(self, strategy: str = 'append') -> Tuple[pd.DataFrame, Dict]:
        """
        Merge dataframes into master table
        strategy: 'append' (stack rows) or 'merge' (join columns by period)
        """
        if not self.data_frames:
            return pd.DataFrame(), {"error": "No data loaded"}
        
        unified_dfs = []
        unification_log = {"merged": [], "conflicts": []}
        
        for name, df in self.data_frames.items():
            date_col = self.detect_date_column(df)
            if date_col:
                df = df.rename(columns={date_col: 'period'})
                
                # Smart date handling:
                # - Excel date serials (e.g., 45292) → convert to dates
                # - Years (2010, 2011) → keep as integers
                # - String dates → keep as strings (no time added)
                is_numeric = pd.api.types.is_numeric_dtype(df['period'])
                
                if is_numeric:
                    min_val = df['period'].min()
                    max_val = df['period'].max()
                    
                    # Excel date serial numbers (typically 25000-60000 for 1968-2064)
                    if min_val > 25000 and max_val < 60000:
                        # Convert Excel serial dates to actual dates
                        df['period'] = pd.to_datetime(df['period'], unit='D', origin='1899-12-30')
                        # Format as date-only strings (no time)
                        df['period'] = df['period'].dt.strftime('%Y-%m-%d')
                    # Years (1900-2100) - keep as integers
                    elif min_val > 1900 and max_val < 2100:
                        pass  # Keep as integers
                
                df['source_file'] = name
                unified_dfs.append(df)
                unification_log["merged"].append(name)
            else:
                unification_log["conflicts"].append(f"{name}: No date column detected")
        
        if unified_dfs:
            if strategy == 'merge':
                # Merge by period (join columns)
                from functools import reduce
                try:
                    # Drop source_file for merging to avoid duplication/conflicts
                    dfs_for_merge = [d.drop(columns=['source_file'], errors='ignore') for d in unified_dfs]
                    
                    self.master_df = reduce(
                        lambda left, right: pd.merge(left, right, on='period', how='outer'), 
                        dfs_for_merge
                    )
                    self.master_df = self.master_df.sort_values('period')
                except Exception as e:
                    unification_log["conflicts"].append(f"Merge failed: {str(e)}")
                    # Fallback to append
                    self.master_df = pd.concat(unified_dfs, ignore_index=True, sort=False)
            else:
                # Default: Append (stack rows)
                self.master_df = pd.concat(unified_dfs, ignore_index=True, sort=False)
                self.master_df = self.master_df.sort_values('period')
            
            # Fill NaNs with None for cleaner display if needed, or keep NaNs for math
            # Keeping NaNs is better for pandas operations
        
        return self.master_df, unification_log
    
    
    def get_available_metrics(self) -> List[str]:
        """Return list of available numeric metrics"""
        if self.master_df is None:
            return []
        
        numeric_cols = self.master_df.select_dtypes(include=[np.number]).columns.tolist()
        return [col for col in numeric_cols if col not in ['period']]
    
    def detect_common_columns(self, sheet_names: List[str] = None) -> List[str]:
        """Detect columns that exist in all selected sheets"""
        if not self.data_frames:
            return []
        
        # If no specific sheets selected, use all loaded sheets
        if sheet_names is None:
            sheet_names = list(self.data_frames.keys())
        
        if len(sheet_names) == 0:
            return []
        
        # Start with columns from first sheet
        common_cols = set(self.data_frames[sheet_names[0]].columns)
        
        # Intersect with each subsequent sheet
        for sheet_name in sheet_names[1:]:
            if sheet_name in self.data_frames:
                common_cols &= set(self.data_frames[sheet_name].columns)
        
        return sorted(list(common_cols))
    
    def get_join_candidates(self, sheet_names: List[str] = None) -> List[str]:
        """Return columns suitable for joining (period-like or unique identifiers)"""
        common_cols = self.detect_common_columns(sheet_names)
        
        join_candidates = []
        
        for col in common_cols:
            # Check if it's a date-like column name
            if any(date_word in str(col).lower() for date_word in self.COMMON_DATE_COLS):
                join_candidates.append(col)
                continue
            
            # Check if values look unique enough to be a key (>80% unique values)
            if sheet_names is None:
                sheet_names = list(self.data_frames.keys())
            
            for sheet_name in sheet_names:
                if sheet_name in self.data_frames:
                    df = self.data_frames[sheet_name]
                    if col in df.columns:
                        uniqueness = df[col].nunique() / len(df) if len(df) > 0 else 0
                        if uniqueness > 0.8:  # 80% unique values
                            if col not in join_candidates:
                                join_candidates.append(col)
                            break
        
        return join_candidates
    
    def preview_join(self, sheet_names: List[str], join_key: str, max_rows: int = 5) -> Dict[str, Any]:
        """Preview what a column-based join would look like"""
        if not sheet_names or len(sheet_names) < 2:
            return {"error": "Need at least 2 sheets to join"}
        
        # Validate join key exists in all sheets
        for sheet_name in sheet_names:
            if sheet_name not in self.data_frames:
                return {"error": f"Sheet '{sheet_name}' not found"}
            if join_key not in self.data_frames[sheet_name].columns:
                return {"error": f"Column '{join_key}' not found in sheet '{sheet_name}'"}
        
        # Perform preview join
        try:
            from functools import reduce
            selected_dfs = [self.data_frames[name] for name in sheet_names]
            
            preview_df = reduce(
                lambda left, right: pd.merge(left, right, on=join_key, how='outer'),
                selected_dfs
            )
            
            return {
                "preview": preview_df.head(max_rows),
                "total_rows": len(preview_df),
                "total_columns": len(preview_df.columns),
                "columns": list(preview_df.columns)
            }
        except Exception as e:
            return {"error": f"Join failed: {str(e)}"}
    
    def preview_append(self, sheet_names: List[str], align_columns: List[str] = None, 
                      include_non_common: bool = False, max_rows: int = 5) -> Dict[str, Any]:
        """Preview what a row-based append would look like"""
        if not sheet_names:
            return {"error": "No sheets selected"}
        
        # Validate sheets exist
        for sheet_name in sheet_names:
            if sheet_name not in self.data_frames:
                return {"error": f"Sheet '{sheet_name}' not found"}
        
        try:
            selected_dfs = []
            
            if align_columns:
                # Use only specified columns
                for sheet_name in sheet_names:
                    df = self.data_frames[sheet_name]
                    
                    if include_non_common:
                        # Keep all columns, but align specified ones
                        selected_dfs.append(df)
                    else:
                        # Only keep aligned columns
                        available_cols = [c for c in align_columns if c in df.columns]
                        selected_dfs.append(df[available_cols])
            else:
                # Use all columns from each sheet
                selected_dfs = [self.data_frames[name] for name in sheet_names]
            
            preview_df = pd.concat(selected_dfs, ignore_index=True, sort=False)
            
            return {
                "preview": preview_df.head(max_rows),
                "total_rows": len(preview_df),
                "total_columns": len(preview_df.columns),
                "columns": list(preview_df.columns),
                "has_nulls": preview_df.isnull().any().any()
            }
        except Exception as e:
            return {"error": f"Append failed: {str(e)}"}

    
    def get_metric_data(self, metric: str) -> pd.Series:
        """Get time series for a specific metric"""
        if self.master_df is None or metric not in self.master_df.columns:
            return pd.Series()
        
        return self.master_df.groupby('period')[metric].sum()


class ForecastEngine:
    """Deterministic forecast computation engine"""
    
    def __init__(self):
        self.assumptions: List[Assumption] = []
        self.events: List[Event] = []
        self.growth_regimes: List[GrowthRegime] = []  # v2: Piecewise growth
        self.base_data: Optional[pd.DataFrame] = None
        self.contribution_breakdown: Dict[str, pd.DataFrame] = {}
        self.scenario_attribution: Dict[str, Dict[str, float]] = {}  # v2: Per-driver attribution
        self.periods_per_year: Optional[int] = None  # For annualized growth rate conversion
        
    def set_base_data(self, data: pd.DataFrame):
        """Set historical data for forecasting"""
        self.base_data = data.copy()
        # Ensure 'period' is the index if it exists as a column
        if 'period' in self.base_data.columns:
            self.base_data = self.base_data.set_index('period')
        
        # Auto-detect data frequency
        self.periods_per_year = self.detect_data_frequency()
    
    def detect_data_frequency(self) -> int:
        """
        Detect data frequency (periods per year) from historical data.
        Returns: 1 (yearly), 4 (quarterly), 12 (monthly), 52 (weekly), 252 (daily)
        """
        if self.base_data is None or len(self.base_data) < 2:
            return 4  # Default to quarterly if insufficient data
        
        index = self.base_data.index
        
        # Method 1: Datetime-based detection
        if pd.api.types.is_datetime64_any_dtype(index):
            try:
                # Calculate median difference between periods
                time_diffs = pd.Series(index).diff().dropna()
                median_diff = time_diffs.median()
                
                # Map to periods per year
                days = median_diff.days
                if days < 2:
                    return 252  # Daily (business days)
                elif days < 10:
                    return 52   # Weekly
                elif days < 45:
                    return 12   # Monthly
                elif days < 200:
                    return 4    # Quarterly
                else:
                    return 1    # Yearly
            except:
                pass
        
        # Method 2: Label-based detection (Q1, Q2, Jan, Feb, etc.)
        sample_indices = [str(idx).upper() for idx in index[:min(20, len(index))]]
        
        # Check for quarterly pattern
        if any('Q' in str(idx) and any(c in '1234' for c in str(idx)) for idx in sample_indices):
            return 4
        
        # Check for month names
        month_names = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 
                      'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC',
                      'JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 
                      'JUNE', 'JULY', 'AUGUST', 'SEPTEMBER', 
                      'OCTOBER', 'NOVEMBER', 'DECEMBER']
        if any(any(month in str(idx).upper() for month in month_names) for idx in sample_indices):
            return 12
        
        # Check for week patterns
        if any('WEEK' in str(idx).upper() or 'WK' in str(idx).upper() for idx in sample_indices):
            return 52
        
        # Method 3: Infer from data length
        data_len = len(self.base_data)
        if data_len >= 100:
            # Likely weekly (2 years of weekly data = 104 periods)
            return 52
        elif data_len >= 40:
            # Likely monthly (3+ years of monthly data)
            return 12
        elif data_len >= 12:
            # Likely quarterly (3+ years of quarterly data)
            return 4
        elif data_len >= 3:
            # Likely yearly
            return 1
        
        # Default to quarterly
        return 4
    
    def add_assumption(self, assumption: Assumption):
        """Add an assumption to the model"""
        self.assumptions.append(assumption)
    
    def add_event(self, event: Event):
        """Add a discrete event"""
        self.events.append(event)

    def remove_assumption(self, assumption_id: str):
        """Remove an assumption by ID"""
        self.assumptions = [a for a in self.assumptions if a.id != assumption_id]
        
    def remove_event(self, event_id: str):
        """Remove an event by ID"""
        self.events = [e for e in self.events if e.id != event_id]
    
    # Safety & Validation Methods
    def validate_assumption(self, assumption: Assumption) -> Dict[str, List[str]]:
        """Validate assumption and return warnings/errors"""
        warnings = []
        errors = []
        
        # Check for duplicates
        for existing in self.assumptions:
            if (existing.metric == assumption.metric and 
                existing.name == assumption.name and
                existing.id != assumption.id):
                errors.append(f"Duplicate assumption '{assumption.name}' already exists for {assumption.metric}")
        
        # Check for range consistency
        if assumption.use_range:
            if assumption.min_value is None or assumption.max_value is None:
                errors.append("Range-based assumption missing min or max value")
            elif assumption.min_value >= assumption.max_value:
                errors.append(f"Invalid range: min ({assumption.min_value:.1%}) must be < max ({assumption.max_value:.1%})")
            elif not (assumption.min_value <= assumption.value <= assumption.max_value):
                warnings.append(f"Base value ({assumption.value:.1%}) outside range [{assumption.min_value:.1%}, {assumption.max_value:.1%}]")
        
        # Warn on extreme growth rates
        if assumption.type == AssumptionType.GROWTH:
            if abs(assumption.value) > 0.5:  # 50%+
                warnings.append(f"High growth rate ({assumption.value:.1%}) - verify this is intentional")
        
        # Check for circular dependencies (simple detection)
        if assumption.dependency_id:
            if self._has_circular_dependency(assumption.id, assumption.dependency_id):
                errors.append(f"Circular dependency detected with {assumption.dependency_id}")
        
        return {'warnings': warnings, 'errors': errors}
    
    def _has_circular_dependency(self, start_id: str, current_id: str, visited: set = None) -> bool:
        """Simple circular dependency detection"""
        if visited is None:
            visited = set()
        
        if current_id == start_id:
            return True
        
        if current_id in visited:
            return False
        
        visited.add(current_id)
        
        # Check if current_id has dependencies
        for assumption in self.assumptions:
            if assumption.id == current_id and assumption.dependency_id:
                if self._has_circular_dependency(start_id, assumption.dependency_id, visited):
                    return True
        
        for event in self.events:
            if event.id == current_id:
                return False  # Events don't have dependencies
        
        return False
    
    def validate_event(self, event: Event, forecast_start: Optional[pd.Timestamp] = None, 
                      forecast_end: Optional[pd.Timestamp] = None) -> Dict[str, List[str]]:
        """Validate event dates and parameters"""
        warnings = []
        errors = []
        
        # Validate based on type (Date vs Int)
        is_date_comparison = False
        if forecast_start is not None and isinstance(forecast_start, (pd.Timestamp, datetime)):
            is_date_comparison = True
        
        try:
            if is_date_comparison:
                # Try datetime conversion
                try:
                    event_date = pd.to_datetime(event.date)
                    if forecast_start and event_date < forecast_start:
                        warnings.append(f"Event '{event.name}' ({event.date}) is before forecast start ({forecast_start.strftime('%Y-%m-%d')})")
                    elif forecast_end and event_date > forecast_end:
                        warnings.append(f"Event '{event.name}' ({event.date}) is after forecast end ({forecast_end.strftime('%Y-%m-%d')})")
                except (ValueError, TypeError):
                    # Could be non-datetime format (e.g., "Q4 2023") - allow through
                    pass
            else:
                # Try numeric comparison
                try:
                    event_period = float(event.date)
                    if forecast_start is not None and event_period < float(forecast_start):
                        warnings.append(f"Event '{event.name}' (Period {event.date}) is before forecast start ({forecast_start})")
                    elif forecast_end is not None and event_period > float(forecast_end):
                        warnings.append(f"Event '{event.name}' (Period {event.date}) is after forecast end ({forecast_end})")
                except (ValueError, TypeError):
                    # Could be string format (e.g., "Q4 2023", "Period-5") - allow through
                    pass
        except Exception:
             # Generic fallback - allow any format
             pass
        
        # Validate impact
        if event.impact_multiplier < 0:
            errors.append(f"Event impact multiplier must be positive (got {event.impact_multiplier})")
        elif abs(event.impact_multiplier - 1.0) > 2.0:  # >200% change
            warnings.append(f"Large event impact ({(event.impact_multiplier - 1) * 100:+.0f}%) - verify this is correct")
        
        return {'warnings': warnings, 'errors': errors}
    
    def detect_assumption_conflicts(self) -> List[Dict[str, Any]]:
        """Detect conflicting assumptions on the same metric"""
        conflicts = []
        
        # Group by metric
        by_metric = {}
        for assumption in self.assumptions:
            if assumption.metric not in by_metric:
                by_metric[assumption.metric] = []
            by_metric[assumption.metric].append(assumption)
        
        # Check for conflicts within each metric
        for metric, assumptions in by_metric.items():
            # Multiple growth assumptions on same metric at same layer
            growth_by_layer = {}
            for a in assumptions:
                if a.type == AssumptionType.GROWTH:
                    layer = a.layer
                    if layer not in growth_by_layer:
                        growth_by_layer[layer] = []
                    growth_by_layer[layer].append(a)
            
            for layer, layer_assumptions in growth_by_layer.items():
                if len(layer_assumptions) > 1:
                    conflicts.append({
                        'type': 'duplicate_growth',
                        'metric': metric,
                        'layer': layer,
                        'assumptions': [a.name for a in layer_assumptions],
                        'message': f"Multiple growth assumptions at layer {layer} for {metric}: {', '.join(a.name for a in layer_assumptions)}"
                    })
        
        return conflicts
    
    def should_apply_assumption(self, assumption: Assumption, period: pd.Timestamp) -> bool:
        """Check if assumption should apply based on dependencies"""
        if assumption.depends_on is None or assumption.apply_if == "always":
            return True
        
        # Find the dependency (event or another assumption)
        dependency_event = None
        for event in self.events:
            if event.id == assumption.depends_on:
                dependency_event = event
                break
        
        if dependency_event is None:
            return True  # Dependency not found, apply anyway
        
        event_date = pd.to_datetime(dependency_event.date)
        
        # Check condition
        if assumption.apply_if == "if_occurs":
            return period >= event_date  # Apply after event occurs
        elif assumption.apply_if == "if_not_occurs":
            return period < event_date  # Apply only before event
        
        return True
    
    def calculate_cagr(self, start_value: float, end_value: float, periods: int) -> float:
        """Calculate compound annual growth rate"""
        if start_value <= 0 or end_value <= 0 or periods <= 0:
            return 0.0
        return (end_value / start_value) ** (1 / periods) - 1
    
    # v2: Growth Regime Methods
    def add_growth_regime(self, regime: GrowthRegime):
        """Add a growth regime (analyst-defined or detected)"""
        self.growth_regimes.append(regime)
    
    def detect_growth_regimes(self, metric: str, min_period_length: int = 3) -> List[GrowthRegime]:
        """Detect piecewise growth regimes from historical data"""
        if self.base_data is None or metric not in self.base_data.columns:
            return []
        
        historical = self.base_data[[metric]].dropna()
        if len(historical) < min_period_length * 2:
            return []  # Not enough data for multiple regimes
        
        # Calculate period-over-period growth rates
        growth_rates = historical[metric].pct_change().dropna()
        
        # Simple heuristic: detect regime changes when growth rate shifts by >30%
        regimes = []
        current_regime_start = historical.index[0]
        current_growth_values = []
        
        for i, (period, growth) in enumerate(growth_rates.items()):
            current_growth_values.append(growth)
            
            # Check if we've accumulated enough data and if there's a regime shift
            if len(current_growth_values) >= min_period_length:
                avg_current = np.mean(current_growth_values)
                
                # Look ahead to see if growth pattern changes
                if i < len(growth_rates) - 1:
                    next_growth = growth_rates.iloc[i + 1] if i + 1 < len(growth_rates) else growth
                    
                    # Regime shift detected
                    if abs(next_growth - avg_current) > 0.3 * abs(avg_current):
                        regime = GrowthRegime(
                            id=f"regime_{metric}_{len(regimes)}",
                            metric=metric,
                            start_date=current_regime_start.strftime('%Y-%m-%d'),
                            end_date=period.strftime('%Y-%m-%d'),
                            growth_rate=avg_current,
                            name=f"Regime {len(regimes) + 1}: {avg_current:.1%} growth",
                            regime_type="detected",
                            notes=f"Detected from {len(current_growth_values)} periods"
                        )
                        regimes.append(regime)
                        current_regime_start = period
                        current_growth_values = []
        
        # Add final regime
        if current_growth_values:
            avg_final = np.mean(current_growth_values)
            regime = GrowthRegime(
                id=f"regime_{metric}_{len(regimes)}",
                metric=metric,
                start_date=current_regime_start.strftime('%Y-%m-%d'),
                end_date=None,  # Ongoing
                growth_rate=avg_final,
                name=f"Current Regime: {avg_final:.1%} growth",
                regime_type="detected",
                notes=f"Latest regime from {len(current_growth_values)} periods"
            )
            regimes.append(regime)
        
        return regimes
    
    def get_applicable_regime(self, metric: str, forecast_date: pd.Timestamp) -> Optional[GrowthRegime]:
        """Get the growth regime applicable for a given forecast date"""
        applicable_regimes = [
            r for r in self.growth_regimes 
            if r.metric == metric and pd.to_datetime(r.start_date) <= forecast_date
        ]
        
        if not applicable_regimes:
            return None
        
        # Return the most recent regime
        return max(applicable_regimes, key=lambda r: pd.to_datetime(r.start_date))

    
    def generate_base_forecast(self, metric: str, periods: int = 5) -> pd.DataFrame:
        """Generate base case forecast using CAGR or growth assumptions"""
        if self.base_data is None or metric not in self.base_data.columns:
            return pd.DataFrame()
        
        historical = self.base_data[[metric]].dropna()
        if len(historical) < 2:
            return pd.DataFrame()
        
        # Get last historical value and date
        last_value = historical[metric].iloc[-1]
        last_date = historical.index[-1]
        
        # Find applicable growth assumption
        growth_rate = None
        growth_source = 'cagr'
        for assumption in self.assumptions:
            if assumption.metric == metric and assumption.type == AssumptionType.GROWTH:
                annual_rate = assumption.value  # This is always an annual rate now
                # Convert annual rate to period rate based on data frequency
                periods_per_year = self.periods_per_year if self.periods_per_year else 4
                growth_rate = (1 + annual_rate) ** (1 / periods_per_year) - 1
                growth_source = f'assumption_{assumption.id}'
                break
        
        # Default to CAGR if no assumption provided
        if growth_rate is None:
            first_value = historical[metric].iloc[0]
            growth_rate = self.calculate_cagr(first_value, last_value, len(historical) - 1)
        
        # Generate forecast with incremental contribution tracking
        if pd.api.types.is_datetime64_any_dtype(historical.index):
            # DateTime Logic
            # Attempt to infer frequency
            freq = pd.infer_freq(historical.index)
            
            if not freq and len(historical) > 1:
                # Fallback heuristics based on average delta
                delta_days = (historical.index[-1] - historical.index[0]).days / (len(historical) - 1)
                
                if 360 <= delta_days <= 370:
                    freq = 'Y'  # Annual
                elif 88 <= delta_days <= 94:
                    freq = 'QS' # Quarterly Start
                elif 28 <= delta_days <= 32:
                    freq = 'MS' # Monthly Start
                elif 6 <= delta_days <= 8:
                    freq = 'W'  # Weekly
                elif 0.9 <= delta_days <= 1.1:
                    freq = 'D'  # Daily
            
            if freq:
                # Standard frequency found
                forecast_dates = pd.date_range(start=last_date, periods=periods + 1, freq=freq)[1:]
            else:
                # AMBIGUOUS / IRREGULAR: Use Average Delta Projection
                # "Period 1, 2, 3" style logic but preserving date types logic
                avg_delta = (historical.index[-1] - historical.index[0]) / (len(historical) - 1)
                forecast_dates = [last_date + (avg_delta * i) for i in range(1, periods + 1)]
                forecast_dates = pd.DatetimeIndex(forecast_dates)
        else:
            # Smart Pattern Logic
            # Detect pattern from historical index/period column
            hist_periods = historical.index.tolist()
            forecast_dates = self.extrapolate_periods(hist_periods, periods)



        # Use shared helper for consistency
        forecast_df = self._calculate_forecast_series(
            start_value=last_value,
            growth_rate=growth_rate,
            periods=periods,
            metric=metric,
            forecast_dates=forecast_dates
        )
        
        # Store contribution breakdown
        self.contribution_breakdown[metric] = forecast_df.copy()
        
        return forecast_df

    def _calculate_forecast_series(self, start_value: float, growth_rate: float, periods: int, metric: str, forecast_dates: List[Any]) -> pd.DataFrame:
        """
        Shared logic for generating a forecast series with event application.
        Ensures Base, Upside, and Downside scenarios all respect events consistentely.
        """
        forecast_values = []
        base_contributions = [] # This will be the pure base value
        event_cols_data = [] # List of dicts
        
        cumulative_event_multiplier = 1.0
        active_permanent_events = []
        active_decay_events = []
        
        for i in range(periods):
            current_step = i + 1
            
            # 1. Base components
            # Note: For compounding, we use (1+r)^t from start
            base_val = start_value * ((1 + growth_rate) ** current_step)
            
            # 2. Check for NEW events triggering at this period
            for event in self.events:
                if event.metric == metric:
                    is_match = False
                    
                    # Get current forecast period value for this step
                    current_period_value = None
                    if hasattr(forecast_dates, '__getitem__') and i < len(forecast_dates):
                        current_period_value = forecast_dates[i]
                    
                    # DEBUG
                    # print(f"DEBUG: Step {i+1}: Checking event '{event.name}' (date='{event.date}') vs period '{current_period_value}'")
                    
                    # Match 1: Direct match against forecast period value (PRIMARY)
                    # This handles years (2025), quarters (Q1 2025), months, etc.
                    if current_period_value is not None:
                        evt_str = str(event.date).strip()
                        
                        if hasattr(current_period_value, 'strftime'):
                            # Period is a datetime - compare date strings
                            curr_str = current_period_value.strftime('%Y-%m-%d')
                            if evt_str == curr_str:
                                is_match = True
                                # print(f"DEBUG: ✅ Matched via datetime: '{evt_str}' == '{curr_str}'")
                        else:
                            # Period is a string/number - normalize and compare
                            period_str = str(current_period_value).strip()
                            # print(f"DEBUG: Comparing '{evt_str}' vs '{period_str}' (types: {type(event.date)} vs {type(current_period_value)})")
                            if evt_str == period_str:
                                is_match = True
                                # print(f"DEBUG: ✅ Matched via string: '{evt_str}' == '{period_str}'")
                    
                    # Match 2: Fallback to step index match (for legacy compatibility)
                    # Only if we haven't matched yet and event.date looks like a step number
                    if not is_match:
                        try:
                            event_as_int = int(float(str(event.date)))
                            if event_as_int == current_step:
                                is_match = True
                                # print(f"DEBUG: ✅ Matched via index: {event_as_int} == {current_step}")
                        except (ValueError, TypeError):
                            pass  # event.date is not a number, skip index matching
                    
                    if is_match:
                        if event.decay_periods == 0:
                            cumulative_event_multiplier *= event.impact_multiplier
                            active_permanent_events.append(event)
                        else:
                            active_decay_events.append({
                                'event': event,
                                'periods_left': event.decay_periods,
                                'initial_periods': event.decay_periods
                            })

            # 3. Calculate Final Value
            # Apply permanent
            val_perm = base_val * cumulative_event_multiplier
            
            # Apply decay
            decay_mult = 1.0
            decay_contribs = {} # id -> delta
            
            next_decay_events = []
            for d in active_decay_events:
                evt = d['event']
                full_delta = evt.impact_multiplier - 1.0
                curr_delta = full_delta * (d['periods_left'] / d['initial_periods'])
                
                decay_mult *= (1.0 + curr_delta)
                
                # Attibution (approx)
                decay_contribs[f"event_{evt.id}"] = val_perm * curr_delta
                
                d['periods_left'] -= 1
                if d['periods_left'] > 0:
                    next_decay_events.append(d)
            active_decay_events = next_decay_events
            
            final_value = val_perm * decay_mult
            
            # 4. Attribution
            # Base
            row_contribs = {'base_contribution': base_val}
            
            # Permanent Events
            perm_delta = val_perm - base_val
            if active_permanent_events:
                bs = [abs(e.impact_multiplier - 1.0) for e in active_permanent_events]
                total_b = sum(bs)
                if total_b != 0:
                    for idx, evt in enumerate(active_permanent_events):
                        share = bs[idx] / total_b
                        row_contribs[f"event_{evt.id}"] = perm_delta * share
            
            # Decay Events
            row_contribs.update(decay_contribs)
            
            event_cols_data.append(row_contribs)
            forecast_values.append(final_value)
            base_contributions.append(base_val)

        # Construct DataFrame
        data_dict = {
            'period': forecast_dates,
            metric: forecast_values,
            'base_contribution': base_contributions, # Pure base level
            'type': 'forecast'
        }
        
        # Flatten event columns
        all_event_keys = set()
        for d in event_cols_data:
            all_event_keys.update(d.keys())
        
        for k in all_event_keys:
            if k == 'base_contribution': continue
            col_vals = [d.get(k, 0.0) for d in event_cols_data]
            data_dict[k] = col_vals
            
        return pd.DataFrame(data_dict)

    def extrapolate_periods(self, history: List[Any], n_steps: int) -> List[Any]:
        """
        Smartly extrapolate the next n_steps periods based on history patterns.
        Supports:
        1. Arithmetic Sequences (2020, 2021 -> 2022)
        2. String-Number Patterns ('Period 1', 'Period 2' -> 'Period 3')
        3. Simple Cycles (Q1..Q4, Jan..Dec)
        4. Fallback: integer sequence continuing from last value
        """
        if not history:
            return list(range(1, n_steps + 1))
            
        last_val = history[-1]
        
        # 1. Try Numeric Extrapolation
        try:
            # Convert to float/int to check for arithmetic progression
            nums = pd.to_numeric(history, errors='coerce')
            if nums.notna().all():
                # Check variance in diffs
                diffs = np.diff(nums)
                if len(diffs) > 0:
                    avg_diff = np.mean(diffs)
                    # If variance is low, assume linear trend
                    if np.std(diffs) < 0.1 * abs(avg_diff) or len(diffs) < 3:
                        # Extrapolate
                        start_num = nums.iloc[-1]
                        step = avg_diff
                        # Round step if it's close to integer
                        if abs(step - round(step)) < 0.01:
                            step = round(step)
                            next_vals = range(int(start_num + step), int(start_num + step * (n_steps + 1)), int(step))
                            return list(next_vals)
                        else:
                            next_vals = [start_num + step * i for i in range(1, n_steps + 1)]
                            return next_vals
        except:
            pass
            
        # 2. Check for Quarter Patterns (BEFORE generic pattern matching!)
        # Quarter logic - handle multiple formats:
        # - "Q1", "Q2", "Q3", "Q4" (simple cycle)
        # - "Q1 2022", "Q2 2022", etc. (with year)
        str_val = str(last_val).strip()
        str_val_upper = str_val.upper()
        
        # Check for quarter pattern (Q1-Q4)
        if 'Q' in str_val_upper:
            # Match "Q1", "Q2", "Q3", "Q4" with optional year
            match_q = re.search(r'Q([1-4])', str_val_upper)
            if match_q:
                curr_q = int(match_q.group(1))
                
                # Check if there's a year component
                year_match = re.search(r'(19|20)\d{2}', str_val)
                if year_match:
                    # Format: "Q1 2022" or similar
                    curr_year = int(year_match.group(0))
                    
                    # Determine the exact format from the input string
                    # Find positions
                    q_pattern = re.search(r'Q[1-4]', str_val_upper)
                    q_str = q_pattern.group()  # "Q1", "Q2", etc.
                    q_start = q_pattern.start()
                    y_start = year_match.start()
                    
                    # Extract separator by looking at what's between Q and year
                    if q_start < y_start:
                        # Q before year: "Q1 2022"
                        separator = str_val[q_start + len(q_str):y_start]
                        q_first = True
                    else:
                        # Year before Q: "2022 Q1"
                        separator = str_val[y_start + 4:q_start]
                        q_first = False
                    
                    next_periods = []
                    for i in range(1, n_steps + 1):
                        # Calculate next quarter and year
                        total_quarters_ahead = curr_q - 1 + i  # Total quarters from start of curr_year
                        next_q = (total_quarters_ahead % 4) + 1
                        years_ahead = total_quarters_ahead // 4
                        next_year = curr_year + years_ahead
                        
                        if q_first:
                            next_periods.append(f"Q{next_q}{separator}{next_year}")
                        else:
                            next_periods.append(f"{next_year}{separator}Q{next_q}")
                    
                    return next_periods
                else:
                    # Simple cycle: "Q1", "Q2", "Q3", "Q4"
                    next_periods = []
                    for i in range(1, n_steps + 1):
                        next_q = ((curr_q + i - 1) % 4) + 1
                        next_periods.append(f"Q{next_q}")
                    return next_periods
            
        # 3. Check for Month Patterns (BEFORE generic pattern matching!)
        # Month logic - handle month names with optional years
        months_short = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
        months_long = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"]
        
        # Extract just the month name (without year) for matching
        month_part = str_val_upper
        year_part = None
        separator = ""
        month_first = True
        
        # Check if there's a year component
        year_match = re.search(r'(19|20)\d{2}', str_val)
        if year_match:
            year_part = year_match.group(0)
            year_pos = year_match.start()
            
            # Determine which part is month vs year
            parts = str_val.split()
            if len(parts) >= 2:
                # Try to find which part is the month
                for i, part in enumerate(parts):
                    if part.upper() in months_short or part.upper() in months_long:
                        month_part = part.upper()
                        if year_match.group(0) in parts[0]:  # Year first
                            month_first = False
                            separator = str_val[year_pos + 4:str_val.upper().index(part.upper())]
                        else:  # Month first
                            month_first = True
                            month_end = str_val.upper().index(part.upper()) + len(part)
                            separator = str_val[month_end:year_pos]
                        break
        
        # Check against month lists
        curr_idx = -1
        is_long = False
        
        if month_part in months_short:
            curr_idx = months_short.index(month_part)
        elif month_part in months_long:
            curr_idx = months_long.index(month_part)
            is_long = True
        
        if curr_idx != -1:
            # Detected month pattern
            next_periods = []
            source_list = months_long if is_long else months_short
            
            # Preserve casing from original input
            casing = "title"
            if str_val.split()[0 if month_first else -1].isupper(): 
                casing = "upper"
            elif str_val.split()[0 if month_first else -1].islower(): 
                casing = "lower"
            
            for i in range(1, n_steps + 1):
                next_idx = (curr_idx + i) % 12
                val = source_list[next_idx]
                
                if casing == "title": 
                    val = val.title()
                elif casing == "lower": 
                    val = val.lower()
                # else upper (default)
                
                if year_part:
                    # Handle year transitions (Dec 2022 → Jan 2023)
                    years_ahead = (curr_idx + i) // 12
                    next_year = int(year_part) + years_ahead
                    
                    if month_first:
                        next_periods.append(f"{val}{separator}{next_year}")
                    else:
                        next_periods.append(f"{next_year}{separator}{val}")
                else:
                    next_periods.append(val)
            
            return next_periods
        
        # 4. Try Pattern Matching (Prefix-Number)
        # Regex to find "Text Number"
        match = re.search(r'([^\d]*)(\d+)([^\d]*)', str(last_val))
        if match:
            prefix, num_str, suffix = match.groups()
            
            # Verify previous value matches same pattern to confirm
            if len(history) > 1:
                prev_match = re.search(r'([^\d]*)(\d+)([^\d]*)', str(history[-2]))
                if prev_match and prev_match.group(1) == prefix and prev_match.group(3) == suffix:
                    # Valid pattern detected
                    current_num = int(num_str)
                    prev_num = int(prev_match.group(2))
                    step = current_num - prev_num
                    
                    next_periods = []
                    for i in range(1, n_steps + 1):
                        next_num = current_num + (step * i)
                        # Padding: preserve leading zeros if present in original (e.g., '01')
                        if num_str.startswith('0') and len(num_str) > 1:
                             fmt_num = f"{next_num:0{len(num_str)}d}"
                        else:
                             fmt_num = str(next_num)
                        next_periods.append(f"{prefix}{fmt_num}{suffix}")
                    return next_periods

        # 5. Fallback: Append integers if not numeric
        try:
             # If last_val was simple int
             return list(range(int(last_val) + 1, int(last_val) + n_steps + 1))
        except:
             # Total fallback: strings
             return [f"Period {i}" for i in range(len(history) + 1, len(history) + n_steps + 1)]
    
    def apply_events(self, forecast_df: pd.DataFrame, metric: str) -> pd.DataFrame:
        """Apply event-based modifications to forecast with contribution tracking"""
        result_df = forecast_df.copy()
        
        # Initialize event contribution columns
        event_contributions = {}
        
        for event in self.events:
            if event.metric != metric:
                continue
            
            event_date = None
            try:
                # Try explicit datetime match first
                event_date = pd.to_datetime(event.date)
            except:
                pass # Might be generic integer

            
            # Track contribution for this event
            event_key = f'event_{event.id}'
            event_contributions[event_key] = [0.0] * len(result_df)
            
            # Find closest forecast period
            # Handle Generic Integer Periods logic
            is_datetime_index = pd.api.types.is_datetime64_any_dtype(result_df['period'])
            
            if not is_datetime_index:
                # Integer Logic (Exact Match)
                try:
                    target_period = int(float(event.date)) # Handle "5" or "5.0"
                    if target_period in result_df['period'].values:
                        idx = result_df[result_df['period'] == target_period].index[0]
                        row_idx = result_df.index.get_loc(idx)
                    else:
                        continue
                except ValueError:
                    continue
            
            # DateTime Logic
            elif event_date in result_df['period'].values:
                idx = result_df[result_df['period'] == event_date].index[0]
                row_idx = result_df.index.get_loc(idx)
            else:
                # Fallback: Match by Year (for annual models)
                # Check if we have a matching year in the forecast
                if pd.api.types.is_datetime64_any_dtype(result_df['period']):
                    matching_years = result_df[result_df['period'].dt.year == event_date.year]
                else:
                    # For string periods, try to extract year
                    matching_years = pd.DataFrame()  # Skip year matching for non-datetime periods
                
                if not matching_years.empty:
                    idx = matching_years.index[0]
                    row_idx = result_df.index.get_loc(idx)
                else:
                    # Event is outside forecast range
                    continue
                
                # Apply multiplier and track contribution
                if event.decay_periods == 0:
                    # Permanent effect
                    for i in range(row_idx, len(result_df)):
                        original = result_df.loc[result_df.index[i], metric]
                        new_value = original * event.impact_multiplier
                        contribution = new_value - original
                        result_df.loc[result_df.index[i], metric] = new_value
                        event_contributions[event_key][i] = contribution
                else:
                    # Decaying effect
                    for i in range(min(event.decay_periods, len(result_df) - row_idx)):
                        decay_factor = 1 - (i / event.decay_periods)
                        multiplier = 1 + (event.impact_multiplier - 1) * decay_factor
                        original = result_df.loc[result_df.index[row_idx + i], metric]
                        new_value = original * multiplier
                        contribution = new_value - original
                        result_df.loc[result_df.index[row_idx + i], metric] = new_value
                        event_contributions[event_key][row_idx + i] = contribution
        
        # Add event contributions to dataframe
        for event_key, contributions in event_contributions.items():
            result_df[event_key] = contributions
        
        return result_df
    
    def generate_scenarios(self, metric: str, periods: int = 5) -> Dict[str, pd.DataFrame]:
        """Generate base, upside, and downside scenarios using consistent event logic."""
        
        # 1. Base Forecast
        # Now uses _calculate_forecast_series internally, so it INCLUDES events
        base_forecast = self.generate_base_forecast(metric, periods)
        
        # v2: Per-driver scenario builder using the event-adjusted base
        upside_forecast = base_forecast.copy()
        downside_forecast = base_forecast.copy()
        
        # Track per-driver contributions
        driver_deltas = {
            'upside': {},
            'downside': {}
        }
        
        # Apply per-driver deltas based on confidence or ranges
        for assumption in self.assumptions:
            if assumption.metric != metric:
                continue
            
            driver_name = assumption.name
            
            if assumption.use_range and assumption.min_value is not None and assumption.max_value is not None:
                # v2: Use explicit ranges
                base_val = assumption.value
                upside_val = assumption.max_value
                downside_val = assumption.min_value
                
                # Calculate delta from base
                upside_delta_pct = (upside_val - base_val) / (1 if base_val == 0 else abs(base_val))
                downside_delta_pct = (downside_val - base_val) / (1 if base_val == 0 else abs(base_val))
                
                # Apply to forecast
                upside_forecast[metric] = upside_forecast[metric] * (1 + upside_delta_pct)
                downside_forecast[metric] = downside_forecast[metric] * (1 + downside_delta_pct)
                
                driver_deltas['upside'][driver_name] = upside_delta_pct * 100
                driver_deltas['downside'][driver_name] = downside_delta_pct * 100
                
            else:
                # v1 fallback: Confidence-based multipliers (but per-driver tracked)
                if assumption.confidence == 'high':
                    # High confidence: small scenario spread (+/-5%)
                    upside_mult = 1.05
                    downside_mult = 0.95
                elif assumption.confidence == 'medium':
                    # Medium confidence: moderate spread (+/-10%)
                    upside_mult = 1.10
                    downside_mult = 0.90
                elif assumption.confidence == 'low':
                    # Low confidence: wide spread (+/-20%)
                    upside_mult = 1.20
                    downside_mult = 0.80
                else:
                    upside_mult = 1.0
                    downside_mult = 1.0
                
                # Settle the delta_pct for legacy path so downstream logic works
                upside_delta_pct = upside_mult - 1.0
                downside_delta_pct = downside_mult - 1.0
                
                if assumption.type == AssumptionType.GROWTH:
                    # Growth assumption: Apply static multiplier to the scenario
                    # NOTE: This is an approximation. Ideally we'd regenerate the forecast
                    # with adjusted growth rates. But a static multiplier is reasonable
                    # for visualizing confidence bounds without exponential explosion.
                    
                    upside_forecast[metric] = upside_forecast[metric] * (1 + upside_delta_pct)
                    downside_forecast[metric] = downside_forecast[metric] * (1 + downside_delta_pct)
                    
                    driver_deltas['upside'][driver_name] = upside_delta_pct * 100
                    driver_deltas['downside'][driver_name] = downside_delta_pct * 100
                    
                else:
                    # Non-growth assumptions (events, pricing, etc.): Apply static scalar shift
                    # This changes the LEVEL
                    upside_forecast[metric] = upside_forecast[metric] * (1 + upside_delta_pct)
                    downside_forecast[metric] = downside_forecast[metric] * (1 + downside_delta_pct)
                    
                    driver_deltas['upside'][driver_name] = (upside_delta_pct) * 100
                    driver_deltas['downside'][driver_name] = (downside_delta_pct) * 100
        
        # Store attribution for later retrieval
        self.scenario_attribution = driver_deltas
        
        return {
            'base': base_forecast,
            'upside': upside_forecast,
            'downside': downside_forecast
        }
    
    def get_scenario_attribution(self) -> Dict[str, Dict[str, float]]:
        """Get per-driver attribution for scenario spread"""
        return getattr(self, 'scenario_attribution', {'upside': {}, 'downside': {}})
    
    def get_contribution_breakdown(self, metric: str, scenario_df: pd.DataFrame) -> pd.DataFrame:
        """Generate detailed contribution breakdown for each forecast period"""
        if metric not in scenario_df.columns:
            return pd.DataFrame()
        
        breakdown_data = []
        
        for idx, row in scenario_df.iterrows():
            period = row['period']
            total = row[metric]
            
            contributions = {
                'period': period,
                'total_value': total,
                'base_growth': row.get('base_contribution', 0)
            }
            
            # Add event contributions
            for col in scenario_df.columns:
                if col.startswith('event_'):
                    event_id = col.replace('event_', '')
                    # Find event name
                    event_name = 'Unknown Event'
                    for event in self.events:
                        if event.id == event_id:
                            event_name = event.name
                            break
                    contributions[event_name] = row.get(col, 0)
            
            # Calculate percentages
            if total != 0:
                # Iterate over a copy of keys to avoid modification during iteration
                for key in list(contributions.keys()):
                    if key not in ['period', 'total_value']:
                        pct_key = f"{key}_pct"
                        contributions[pct_key] = (contributions[key] / total) * 100
            
            breakdown_data.append(contributions)
        
        return pd.DataFrame(breakdown_data)
    
    def validate_data_quality(self, metric: str) -> Dict[str, Any]:
        """Validate data quality and return warnings"""
        warnings = {
            'errors': [],
            'warnings': [],
            'quality_score': 'good'
        }
        
        if self.base_data is None or metric not in self.base_data.columns:
            warnings['errors'].append("Metric not found in data")
            warnings['quality_score'] = 'poor'
            return warnings
        
        data = self.base_data[metric].dropna()
        
        # Check 1: Minimum data points
        if len(data) < 3:
            warnings['errors'].append(f"Insufficient data: only {len(data)} points (need at least 3)")
            warnings['quality_score'] = 'poor'
        
        # Check 2: Time continuity
        if len(data) > 1:
            if pd.api.types.is_datetime64_any_dtype(data.index):
                # DateTime Logic
                date_diffs = data.index.to_series().diff()
                # Check for large irregularity (std dev > 180 days is huge)
                try:
                    std_dev = date_diffs.std()
                    if pd.notnull(std_dev) and std_dev > pd.Timedelta(days=180):
                        warnings['warnings'].append("Inconsistent time intervals detected")
                        if warnings['quality_score'] == 'good':
                            warnings['quality_score'] = 'fair'
                except:
                    pass
            else:
                # Integer/Numeric/String Logic
                # Convert to numeric if possible (handles string years like "2020", "2021")
                try:
                    numeric_index = pd.to_numeric(data.index, errors='coerce')
                    if numeric_index.notna().all():
                        # Successfully converted to numeric
                        diffs = numeric_index.to_series().diff().dropna()
                    else:
                        # Some values couldn't convert, skip interval check
                        diffs = pd.Series(dtype=float)
                except:
                    # If conversion fails, skip interval check
                    diffs = pd.Series(dtype=float)
                
                if len(diffs) > 0:
                    std_dev = diffs.std()
                    mean_step = diffs.mean()
                    # If std deviation is > 50% of the average step size, it's irregular
                    if mean_step > 0 and (std_dev / mean_step) > 0.5:
                         warnings['warnings'].append(f"Irregular period intervals detected (std={std_dev:.1f})")
                         if warnings['quality_score'] == 'good':
                            warnings['quality_score'] = 'fair'
        
        # Check 3: Volatility
        if len(data) > 2:
            pct_changes = data.pct_change().dropna()
            if len(pct_changes) > 0:
                volatility = pct_changes.std()
                if volatility > 0.5:  # 50% standard deviation
                    warnings['warnings'].append(f"High volatility detected (sigma={volatility:.1%})")
                    if warnings['quality_score'] == 'good':
                        warnings['quality_score'] = 'fair'
        
        # Check 4: Missing values
        total_rows = len(self.base_data)
        missing_count = total_rows - len(data)
        if missing_count > total_rows * 0.2:  # >20% missing
            warnings['warnings'].append(f"{missing_count} missing values ({missing_count/total_rows:.0%})")
        
        # Check 5: Negative values (if revenue/volume metric)
        if any(data < 0):
            warnings['warnings'].append("Negative values detected")
        
        return warnings


class ExcelOutputGenerator:
    """Generate final Excel output with all components"""
    
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        self.metadata = {
            'created_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'model_version': '1.0',
            'scenario_definitions': {}
        }
    
    def create_readme_sheet(self, assumptions: List[Assumption], events: List[Event], metrics: List[str]):
        """Create comprehensive README sheet"""
        ws = self.workbook.create_sheet("READ ME", 0)  # First sheet
        
        # Title
        ws['A1'] = 'FORECASTING MODEL - READ ME'
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Model Purpose
        ws[f'A{row}'] = 'MODEL PURPOSE'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = 'Assumption-driven forecasting model for structured, auditable scenario planning.'
        row += 1
        ws[f'A{row}'] = f'Created: {self.metadata["created_date"]}'
        row += 2
        
        # Forecasted Metrics
        ws[f'A{row}'] = 'FORECASTED METRICS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        for metric in metrics:
            ws[f'A{row}'] = f'- {metric}'
            row += 1
        row += 1
        
        # Key Assumptions Summary
        ws[f'A{row}'] = 'KEY ASSUMPTIONS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = 'Assumption'
        ws[f'B{row}'] = 'Value'
        ws[f'C{row}'] = 'Confidence'
        ws[f'D{row}'] = 'Source'
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        row += 1
        
        for assumption in assumptions:
            ws[f'A{row}'] = assumption.name
            ws[f'B{row}'] = f'{assumption.value:.1%}' if assumption.type == AssumptionType.GROWTH else str(assumption.value)
            ws[f'C{row}'] = assumption.confidence.upper()
            ws[f'D{row}'] = assumption.source
            
            # Color code by source
            if assumption.source == 'analyst':
                ws[f'B{row}'].font = Font(color='0000FF')  # Blue for analyst input
            elif assumption.source == 'llm':
                ws[f'B{row}'].font = Font(color='FF6600')  # Orange for LLM
            
            row += 1
        row += 1
        
        # Planned Events
        ws[f'A{row}'] = 'PLANNED EVENTS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = 'Event'
        ws[f'B{row}'] = 'Date'
        ws[f'C{row}'] = 'Impact'
        ws[f'D{row}'] = 'Type'
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        row += 1
        
        for event in events:
            ws[f'A{row}'] = event.name
            ws[f'B{row}'] = event.date
            ws[f'C{row}'] = f'{(event.impact_multiplier - 1) * 100:+.0f}%'
            ws[f'D{row}'] = event.event_type.value
            row += 1
        row += 1
        
        # Scenario Definitions
        ws[f'A{row}'] = 'SCENARIO DEFINITIONS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = 'BASE CASE: Uses all assumptions and events as defined'
        row += 1
        ws[f'A{row}'] = 'UPSIDE: Optimistic scenario weighted by assumption confidence'
        row += 1
        ws[f'A{row}'] = 'DOWNSIDE: Conservative scenario weighted by assumption confidence'
        row += 2
        
        # Color Legend
        ws[f'A{row}'] = 'COLOR CODING LEGEND'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = 'Blue text'
        ws[f'A{row}'].font = Font(color='0000FF')
        ws[f'B{row}'] = 'Analyst inputs / assumptions'
        row += 1
        
        ws[f'A{row}'] = 'Black text'
        ws[f'B{row}'] = 'Formulas and calculations'
        row += 1
        
        ws[f'A{row}'] = 'Orange text'
        ws[f'A{row}'].font = Font(color='FF6600')
        ws[f'B{row}'] = 'LLM-suggested values'
        row += 1
        
        ws[f'A{row}'] = 'Green text'
        ws[f'A{row}'].font = Font(color='008000')
        ws[f'B{row}'] = 'Cross-sheet references'
        row += 1
        
        ws[f'A{row}'] = 'Yellow highlight'
        ws[f'A{row}'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws[f'B{row}'] = 'Key assumptions requiring review'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        return ws
        
    def create_assumptions_sheet(self, assumptions: List[Assumption]):
        """Create assumptions reference sheet"""
        ws = self.workbook.create_sheet("Assumptions")
        
        # Headers
        headers = ['ID', 'Type', 'Metric', 'Value', 'Start Date', 'End Date', 'Confidence', 'Source', 'Notes']
        ws.append(headers)
        
        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Add assumption data
        for assumption in assumptions:
            ws.append([
                assumption.id,
                assumption.type.value,
                assumption.metric,
                assumption.value,
                assumption.start_date or '',
                assumption.end_date or '',
                assumption.confidence,
                assumption.source,
                assumption.notes
            ])
            
            # Color code the value based on source
            row_num = ws.max_row
            value_cell = ws.cell(row_num, 4)
            if assumption.source == 'analyst':
                value_cell.font = Font(color='0000FF')  # Blue for inputs
        
        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
    
    def create_events_sheet(self, events: List[Event]):
        """Create events reference sheet"""
        ws = self.workbook.create_sheet("Events")
        
        headers = ['ID', 'Type', 'Metric', 'Date', 'Impact (%)', 'Decay Periods', 'Notes']
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(1, col_num)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for event in events:
            ws.append([
                event.id,
                event.event_type.value,
                event.metric,
                event.date,
                f"{(event.impact_multiplier - 1) * 100:.1f}%",
                event.decay_periods,
                event.notes
            ])
    
    def create_forecast_sheet(self, metric: str, scenarios: Dict[str, pd.DataFrame]):
        """Create forecast sheet with scenarios"""
        ws = self.workbook.create_sheet(f"Forecast_{metric}")
        
        # Freeze panes at B2 (freeze first column and row)
        ws.freeze_panes = 'B2'
        
        # Headers
        ws.append(['Period', 'Base Case', 'Upside', 'Downside'])
        
        for col_num in range(1, 5):
            cell = ws.cell(1, col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Combine scenarios into one table
        # Handle both datetime and string periods
        if pd.api.types.is_datetime64_any_dtype(scenarios['base']['period']):
            periods = scenarios['base']['period'].dt.strftime('%Y-%m-%d').tolist()
        else:
            periods = scenarios['base']['period'].astype(str).tolist()
        
        for i, period in enumerate(periods):
            row_data = [
                period,
                scenarios['base'][metric].iloc[i],
                scenarios['upside'][metric].iloc[i],
                scenarios['downside'][metric].iloc[i]
            ]
            ws.append(row_data)
            
            # Format numbers
            for col_num in range(2, 5):
                cell = ws.cell(ws.max_row, col_num)
                cell.number_format = '#,##0'
        
        # Auto-width
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)
    
    def create_contribution_sheet(self, metric: str, contribution_df: pd.DataFrame):
        """Create detailed contribution breakdown sheet"""
        ws = self.workbook.create_sheet(f"Contribution_{metric}")
        
        # Freeze panes
        ws.freeze_panes = 'B2'
        
        # Write dataframe
        for r_idx, row in enumerate(dataframe_to_rows(contribution_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Header formatting
                if r_idx == 1:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                # Number formatting for value columns
                if r_idx > 1 and c_idx > 1:
                    if '_pct' in str(contribution_df.columns[c_idx - 1]):
                        cell.number_format = '0.0%'
                    elif 'value' in str(contribution_df.columns[c_idx - 1]) or 'contribution' in str(contribution_df.columns[c_idx - 1]):
                        cell.number_format = '#,##0'
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 25)
    
    def save(self, filepath: Path):
        """Save workbook"""
        self.workbook.save(filepath)
