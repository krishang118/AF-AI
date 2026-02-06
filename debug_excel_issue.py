#!/usr/bin/env python3
"""
DEBUGGING THE EXACT ISSUE
Check the exported Excel file manually
"""

import pandas as pd
import tempfile
import openpyxl
from data_joiner import DataJoiner
from excel_formula_exporter import export_combined_excel_with_formulas

# Create test data
df1 = pd.DataFrame({
    'period': ['2020', '2021', '2022'],
    'Revenue': [100, 110, 121]
})
df2 = pd.DataFrame({
    'period': ['2020', '2021', '2022'],
    'Cost': [40, 45, 50]
})

# Join
joiner = DataJoiner()
combined = joiner.join_on_column({'Sheet1': df1, 'Sheet2': df2}, 'period', mode='reference')

# Export
output_file = '/tmp/debug_export.xlsx'
export_combined_excel_with_formulas(
    joined_df=combined,
    source_mapping=joiner.get_source_mapping(),
    source_dataframes={'Sheet1': df1, 'Sheet2': df2},
    output_path=output_file,
    join_type='column_join'
)

print(f"Exported to: {output_file}")
print("\n" + "="*80)

# Inspect the file
wb = openpyxl.load_workbook(output_file)
print(f"Sheet names: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    print(f"\n\nSheet: {sheet_name}")
    ws = wb[sheet_name]
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                row_data.append(f"FORMULA: {cell.value}")
            else:
                row_data.append(f"VALUE: {cell.value}")
        print(f"  Row {row_idx}: {row_data}")

wb.close()

# Try data-only mode
print("\n\n" + "="*80)
print("Reading with pandas (default - data_only=False):")
try:
    df = pd.read_excel(output_file, sheet_name='Combined_Data')
    print(df)
    print(f"Shape: {df.shape}")
except Exception as e:
    print(f"Error: {e}")

# Try with data_only=True
print("\n\nReading with openpyxl data_only=True:")
try:
    wb_data = openpyxl.load_workbook(output_file, data_only=True)
    ws_data = wb_data['Combined_Data']
    
    # Read manually
    data = []
    headers = [ws_data.cell(1, col).value for col in range(1, ws_data.max_column + 1)]
    for row in range(2, ws_data.max_row + 1):
        row_data = [ws_data.cell(row, col).value for col in range(1, ws_data.max_column + 1)]
        data.append(row_data)
    
    df_manual = pd.DataFrame(data, columns=headers)
    print(df_manual)
    wb_data.close()
except Exception as e:
    print(f"Error: {e}")

print("\n\n" + "="*80)
print(f"\nFile saved at: {output_file}")
print("You can open this file in Excel to verify")
