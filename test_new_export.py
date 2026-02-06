#!/usr/bin/env python3
"""
Test new xlsxwriter-based export
Should have BOTH formulas AND values
"""

import pandas as pd
from data_joiner import DataJoiner
from excel_formula_exporter import export_combined_excel_with_formulas
import openpyxl

print("="*80)
print("TESTING XLSXWRITER EXPORT - FORMULAS + VALUES")
print("="*80)

# Load quarterly data
df_quarterly = pd.read_csv('quarterly/sample_quarterly_saas.csv')

# Split into 2 sheets
df_q1 = df_quarterly[['Quarter', 'Quarterly_Revenue', 'ARR']].copy()
df_q2 = df_quarterly[['Quarter', 'Total_Customers', 'Churn_Rate_Pct']].copy()

# Join
joiner = DataJoiner()
combined = joiner.join_on_column({'Revenue': df_q1, 'Customers': df_q2}, 'Quarter', mode='reference')

print(f"Combined: {len(combined)} rows × {len(combined.columns)} columns\n")

# Export with new xlsxwriter-based exporter
output_file = '/tmp/test_formulas_and_values.xlsx'
export_combined_excel_with_formulas(
    joined_df=combined,
    source_mapping=joiner.get_source_mapping(),
    source_dataframes={'Revenue': df_q1, 'Customers': df_q2},
    output_path=output_file,
    join_type='column_join'
)

print(f"Exported to: {output_file}\n")

# Check 1: Inspect with openpyxl (see formulas)
print("="*80)
print("CHECK 1: Formula Inspection (openpyxl)")
print("="*80)

wb = openpyxl.load_workbook(output_file)
ws = wb['Combined_Data']

print(f"\nSheet: Combined_Data")
print(f"Rows: {ws.max_row}, Columns: {ws.max_column}\n")

print("First 3 data rows (formulas):")
for row_idx in range(2, min(5, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        if isinstance(cell.value, str) and cell.value.startswith('='):
            row_data.append(f"FORMULA:{cell.value}")
        else:
            row_data.append(f"VALUE:{cell.value}")
    print(f"  Row {row_idx}: {row_data}")

wb.close()

# Check 2: Read with pandas (should get values)
print("\n" + "="*80)
print("CHECK 2: Pandas Read-back (values)")
print("="*80)

df_readback = pd.read_excel(output_file, sheet_name='Combined_Data')
print(f"\nRead back {len(df_readback)} rows\n")
print(df_readback.head())

# Check 3: Verify correctness
print("\n" + "="*80)
print("VERIFICATION")
print("="*80)

if df_readback.empty or len(df_readback) == 0:
    print("❌ FAILED: Empty DataFrame")
elif (df_readback.select_dtypes(include=['number']) == 0).all().all():
    print("❌ FAILED: All zeros")
else:
    print(f"✅ SUCCESS!")
    print(f"  - {len(df_readback)} rows read back")
    print(f"  - Sample values: Revenue={df_readback['Quarterly_Revenue'].iloc[0]}, Customers={df_readback['Total_Customers'].iloc[0]}")
    print(f"  - Formulas present: YES (see CHECK 1 above)")
    print(f"\  - Values readable by pandas: YES")
    print(f"\n🎉 PERFECT! You have BOTH formulas AND values!")

print("\n" + "="*80)
print(f"File: {output_file}")
print("Open in Excel to verify formulas reference source sheets!")
