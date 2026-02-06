#!/usr/bin/env python3
"""
Debug the user's downloaded file
"""

import openpyxl
import sys

if len(sys.argv) < 2:
    print("Usage: python debug_user_file.py /path/to/downloaded/file.xlsx")
    print("\nOr just drag-drop your downloaded Excel file here:")
    file_path = input("> ").strip().strip("'\"")
else:
    file_path = sys.argv[1]

print("="*80)
print(f"INSPECTING: {file_path}")
print("="*80)

try:
    wb = openpyxl.load_workbook(file_path)
    
    print(f"\n📊 Sheet names: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'='*80}")
        print(f"SHEET: {sheet_name}")
        print(f"{'='*80}")
        print(f"Rows: {ws.max_row}, Columns: {ws.max_column}\n")
        
        # Check first 5 rows
        formula_count = 0
        value_count = 0
        
        print("First 5 rows:")
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(6, ws.max_column + 1)):  # First 5 columns
                cell = ws.cell(row_idx, col_idx)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    row_data.append(f"FORMULA")
                    formula_count += 1
                elif cell.value is None:
                    row_data.append("NULL")
                else:
                    row_data.append(f"VALUE:{cell.value}")
                    value_count += 1
            print(f"  Row {row_idx}: {row_data}")
        
        print(f"\n  Formula cells: {formula_count}")
        print(f"  Value cells: {value_count}")
        
        if sheet_name == 'Combined_Data' and formula_count == 0:
            print("\n  ❌ NO FORMULAS IN COMBINED_DATA!")
            print("  This means it's using simple to_excel(), not formula export")
    
    wb.close()
    
except Exception as e:
    print(f"\n❌ Error: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "="*80)
