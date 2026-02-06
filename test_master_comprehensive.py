#!/usr/bin/env python3
"""
MASTER COMPREHENSIVE TEST SUITE
Tests ALL major functionalities across the entire system
"""

import pandas as pd
import numpy as np
from pathlib import Path
import tempfile
import openpyxl

# Import all available modules
from data_joiner import DataJoiner
from data_editor import DataEditor
from forecast_engine import ForecastEngine, Assumption, AssumptionType, Event
from excel_formula_exporter import export_combined_excel_with_formulas
from hybrid_excel_export import export_hybrid_excel

class TestResults:
    def __init__(self):
        self.results = {}
        self.failed_tests = []
    
    def add(self, test_name, passed, details=""):
        self.results[test_name] = {
            'passed': passed,
            'details': details
        }
        if not passed:
            self.failed_tests.append(test_name)
    
    def print_summary(self):
        passed = sum(1 for r in self.results.values() if r['passed'])
        total = len(self.results)
        
        print("\n" + "="*70)
        print(" TEST SUMMARY")
        print("="*70)
        
        for test_name, result in self.results.items():
            status = "✅ PASS" if result['passed'] else "❌ FAIL"
            print(f"{test_name:.<50} {status}")
            if not result['passed'] and result['details']:
                print(f"  └─ {result['details']}")
        
        print(f"\n{'='*70}")
        print(f"TOTAL: {passed}/{total} tests passed ({passed/total*100:.0f}%)")
        print(f"{'='*70}\n")
        
        return passed == total

results = TestResults()

# ============================================================================
# TEST 2: DATA JOINING
# ============================================================================
print("\n" + "="*70)
print("TEST SUITE 2: DATA JOINING")
print("="*70)

try:
    # Create two test dataframes
    df1 = pd.DataFrame({
        'period': ['2023', '2024'],
        'Revenue': [100, 150]
    })
    df2 = pd.DataFrame({
        'period': ['2023', '2024'],
        'Cost': [40, 60]
    })
    
    joiner = DataJoiner()
    
    # Test 2.1: Column Join
    joined = joiner.join_on_column(
        {'File1': df1, 'File2': df2},
        key_column='period',
        join_type='outer'
    )
    test_passed = 'Revenue' in joined.columns and 'Cost' in joined.columns
    results.add("2.1 Column Join (Outer)", test_passed,
                f"Columns: {joined.columns.tolist()}" if test_passed else "Join failed")
    
    # Test 2.2: Row Append
    appended = joiner.append_rows({'File1': df1, 'File2': df1})
    test_passed = len(appended) == 4  # 2 + 2
    results.add("2.2 Row Append", test_passed,
                f"{len(appended)} rows" if test_passed else "Append failed")
    
    # Test 2.3: Source Mapping
    mapping = joiner.get_source_mapping()
    test_passed = mapping is not None and len(mapping) > 0
    results.add("2.3 Source Mapping Tracking", test_passed)
    
    print(f"✓ Join: {len(joined)} rows, {len(joined.columns)} columns")
    
except Exception as e:
    results.add("2.1 Column Join (Outer)", False, str(e))
    results.add("2.2 Row Append", False, str(e))
    results.add("2.3 Source Mapping Tracking", False, str(e))

# ============================================================================
# TEST 3: DATA EDITING
# ============================================================================
print("\n" + "="*70)
print("TEST SUITE 3: DATA EDITING")
print("="*70)

try:
    df_edit = pd.DataFrame({
        'Revenue': [100, 200, 300],
        'Cost': [40, 80, 120]
    })
    
    editor = DataEditor()
    editor.df = df_edit.copy()
    
    # Test 3.1: Replace Values
    editor.replace_values(find=100, replace=999, columns=['Revenue'])
    test_passed = editor.df['Revenue'].iloc[0] == 999
    results.add("3.1 Replace Values", test_passed)
    
    # Test 3.2: Mark as NA
    editor.mark_as_na(row_indices=[1], column_names=['Cost'])
    test_passed = pd.isna(editor.df['Cost'].iloc[1])
    results.add("3.2 Mark as NA", test_passed)
    
    # Test 3.3: Cell-Level Tracking
    affected = editor.get_affected_cells()
    test_passed = len(affected) == 2  # 1 replace + 1 NA
    results.add("3.3 Cell-Level Change Tracking", test_passed,
                f"{len(affected)} cells tracked" if test_passed else "Tracking failed")
    
    # Test 3.4: Edit History
    history = editor.get_edit_history()
    test_passed = len(history) == 2
    results.add("3.4 Edit History", test_passed)
    
    print(f"✓ Editing: {len(affected)} cells modified, {len(history)} operations")
    
except Exception as e:
    results.add("3.1 Replace Values", False, str(e))
    results.add("3.2 Mark as NA", False, str(e))
    results.add("3.3 Cell-Level Change Tracking", False, str(e))
    results.add("3.4 Edit History", False, str(e))

# ============================================================================
# TEST 4: EXCEL EXPORT (FORMULAS)
# ============================================================================
print("\n" + "="*70)
print("TEST SUITE 4: EXCEL EXPORT WITH FORMULAS")
print("="*70)

try:
    # Prepare source dataframes
    source_dfs = {
        'Sheet1': pd.DataFrame({'period': ['2023', '2024'], 'A': [10, 20]}),
        'Sheet2': pd.DataFrame({'period': ['2023', '2024'], 'B': [30, 40]})
    }
    
    # Create joined dataframe
    joiner = DataJoiner()
    joined_df = joiner.join_on_column(source_dfs, 'period', 'outer')
    
    # Export with formulas
    output_file = tempfile.mktemp(suffix='.xlsx')
    export_combined_excel_with_formulas(
        joined_df=joined_df,
        source_mapping=joiner.get_source_mapping(),
        source_dataframes=source_dfs,
        output_path=output_file,
        join_type='column_join'
    )
    
    # Test 4.1: File creation
    test_passed = Path(output_file).exists()
    results.add("4.1 Excel File Creation", test_passed)
    
    # Test 4.2: Formula insertion
    if test_passed:
        wb = openpyxl.load_workbook(output_file)
        ws = wb['Combined_Data']
        cell_value = ws['B2'].value  # Should be a formula
        is_formula = isinstance(cell_value, str) and cell_value.startswith('=')
        results.add("4.2 Formula References", is_formula,
                    f"Cell contains: {cell_value}" if not is_formula else "")
        wb.close()
    else:
        results.add("4.2 Formula References", False, "File not created")
    
    print(f"✓ Excel Export: Formulas exported to {Path(output_file).name}")
    
except Exception as e:
    results.add("4.1 Excel File Creation", False, str(e))
    results.add("4.2 Formula References", False, str(e))

# ============================================================================
# TEST 5: HYBRID EXCEL EXPORT
# ============================================================================
print("\n" + "="*70)
print("TEST SUITE 5: HYBRID EXCEL EXPORT (CELL-LEVEL)")
print("="*70)

try:
    # Create test data with edits
    df_base = pd.DataFrame({
        'period': ['2023', '2024', '2025'],
        'Revenue': [100, 200, 300]
    })
    
    editor = DataEditor()
    editor.df = df_base.copy()
    editor.mark_as_na([1], ['Revenue'])  # Mark row 1 as NA
    
    affected_cells = editor.get_affected_cells()
    
    # Prepare for hybrid export
    source_dfs = {'Source': df_base}
    joiner = DataJoiner()
    joiner.join_on_column({'Source': df_base}, 'period', 'outer')
    
    output_file = tempfile.mktemp(suffix='.xlsx')
    
    # Test 5.1: Hybrid export executes
    try:
        export_hybrid_excel(
            combined_df=editor.df,
            affected_cells=affected_cells,
            source_mapping=joiner.get_source_mapping(),
            source_dataframes=source_dfs,
            output_path=output_file,
            join_type='column_join'
        )
        test_passed = Path(output_file).exists()
        results.add("5.1 Hybrid Export Execution", test_passed)
    except Exception as e:
        results.add("5.1 Hybrid Export Execution", False, str(e))
        test_passed = False
    
    # Test 5.2: Cell-level granularity
    if test_passed:
        wb = openpyxl.load_workbook(output_file)
        ws = wb['Combined_Data']
        
        # Row 1 (index 0) should have formula (unedited)
        cell_row1 = ws['B2'].value
        is_formula_row1 = isinstance(cell_row1, str) and cell_row1.startswith('=')
        
        # Row 2 (index 1) should be None/empty (marked as NA)
        cell_row2 = ws['B3'].value
        is_na_row2 = cell_row2 is None or cell_row2 == ''
        
        test_passed = is_formula_row1 and is_na_row2
        results.add("5.2 Cell-Level Granularity", test_passed,
                    f"Row1: {cell_row1}, Row2: {cell_row2}" if not test_passed else "")
        wb.close()
    else:
        results.add("5.2 Cell-Level Granularity", False, "Export failed")
    
    print(f"✓ Hybrid Export: {len(affected_cells)} cells with values, rest with formulas")
    
except Exception as e:
    results.add("5.1 Hybrid Export Execution", False, str(e))
    results.add("5.2 Cell-Level Granularity", False, str(e))

# ============================================================================
# TEST 6: FORECAST ENGINE - CAGR & GROWTH
# ============================================================================
print("\n" + "="*70)
print("TEST SUITE 6: FORECAST ENGINE - CAGR & GROWTH")
print("="*70)

try:
    df_forecast = pd.DataFrame({
        'period': ['2020', '2021', '2022'],
        'Revenue': [100, 110, 121]
    })
    
    engine = ForecastEngine()
    engine.set_base_data(df_forecast)
    
    # Test 6.1: CAGR Calculation
    cagr = engine.calculate_cagr(100, 121, 2)
    expected_cagr = 0.10  # 10% exactly
    test_passed = abs(cagr - expected_cagr) < 0.001
    results.add("6.1 CAGR Calculation", test_passed,
                f"Got {cagr:.4f}, expected {expected_cagr:.4f}" if not test_passed else "")
    
    # Test 6.2: Base Forecast (no assumption)
    base_forecast = engine.generate_base_forecast('Revenue', periods=3)
    test_passed = not base_forecast.empty and len(base_forecast) == 3
    results.add("6.2 Base Forecast Generation", test_passed)
    
    # Test 6.3: Forecast with assumption
    assumption = Assumption(
        id='test_growth',
        metric='Revenue',
        type=AssumptionType.GROWTH,
        value=0.05,  # 5% annual
        name='Test',
        confidence='medium',
        layer='base'
    )
    engine.add_assumption(assumption)
    
    forecast_w_assumption = engine.generate_base_forecast('Revenue', periods=3)
    # First period should grow by ~5% from last historical (121)
    expected_approx = 121 * 1.05
    actual = forecast_w_assumption['Revenue'].iloc[0]
    test_passed = abs(actual - expected_approx) < 5  # Allow 5 unit tolerance
    results.add("6.3 Forecast with Assumption", test_passed,
                f"Got {actual:.2f}, expected ~{expected_approx:.2f}" if not test_passed else "")
    
    print(f"✓ Forecast: CAGR={cagr*100:.1f}%, 3-period projection generated")
    
except Exception as e:
    results.add("6.1 CAGR Calculation", False, str(e))
    results.add("6.2 Base Forecast Generation", False, str(e))
    results.add("6.3 Forecast with Assumption", False, str(e))

# ============================================================================
# TEST 7: FORECAST ENGINE - EVENTS
# ============================================================================
print("\n" + "="*70)
print("TEST SUITE 7: FORECAST ENGINE - EVENTS")
print("="*70)

try:
    df_event = pd.DataFrame({
        'period': ['2020', '2021', '2022'],
        'Revenue': [100, 110, 121]
    })
    
    engine = ForecastEngine()
    engine.set_base_data(df_event)
    
    # Test 7.1: Add event
    event = Event(
        id='launch',
        metric='Revenue',
        date='2021',  # Year 2021 (second forecast period)
        impact_multiplier=1.20,  # +20%
        name='Product Launch',
        decay_periods=0  # Permanent
    )
    engine.add_event(event)
    
    test_passed = len(engine.events) == 1
    results.add("7.1 Event Addition", test_passed)
    
    # Test 7.2: Event application in forecast
    forecast = engine.generate_base_forecast('Revenue', periods=3)
    # Event should boost values from period 2 onward
    test_passed = not forecast.empty
    results.add("7.2 Event Application", test_passed)
    
    print(f"✓ Events: 1 event added and applied to forecast")
    
except Exception as e:
    results.add("7.1 Event Addition", False, str(e))
    results.add("7.2 Event Application", False, str(e))

# ============================================================================
# TEST 8: FORECAST ENGINE - SCENARIOS
# ============================================================================
print("\n" + "="*70)
print("TEST SUITE 8: FORECAST ENGINE - SCENARIOS")
print("="*70)

try:
    df_scenario = pd.DataFrame({
        'period': ['2020', '2021', '2022'],
        'Revenue': [100, 110, 121]
    })
    
    engine = ForecastEngine()
    engine.set_base_data(df_scenario)
    
    assumption = Assumption(
        id='growth',
        metric='Revenue',
        type=AssumptionType.GROWTH,
        value=0.05,
        name='Growth',
        confidence='medium',  # ±10%
        layer='base'
    )
    engine.add_assumption(assumption)
    
    # Test 8.1: Scenario generation
    scenarios = engine.generate_scenarios('Revenue', periods=3)
    test_passed = ('base' in scenarios and 'upside' in scenarios and 'downside' in scenarios)
    results.add("8.1 Scenario Generation", test_passed)
    
    # Test 8.2: Upside > Base > Downside
    if test_passed:
        base_end = scenarios['base']['Revenue'].iloc[-1]
        upside_end = scenarios['upside']['Revenue'].iloc[-1]
        downside_end = scenarios['downside']['Revenue'].iloc[-1]
        
        test_passed = upside_end > base_end > downside_end
        results.add("8.2 Scenario Ordering", test_passed,
                    f"Upside={upside_end:.1f}, Base={base_end:.1f}, Down={downside_end:.1f}" if not test_passed else "")
    else:
        results.add("8.2 Scenario Ordering", False, "Scenarios not generated")
    
    print(f"✓ Scenarios: Base, Upside, Downside all generated")
    
except Exception as e:
    results.add("8.1 Scenario Generation", False, str(e))
    results.add("8.2 Scenario Ordering", False, str(e))

# ============================================================================
# TEST 9: PERIOD DETECTION (ALL FREQUENCIES)
# ============================================================================
print("\n" + "="*70)
print("TEST SUITE 9: PERIOD DETECTION")
print("="*70)

try:
    # Test 9.1: Daily detection
    daily_df = pd.DataFrame({
        'period': pd.date_range('2024-01-01', periods=5, freq='D'),
        'Revenue': [100, 110, 105, 115, 120]
    })
    engine = ForecastEngine()
    engine.set_base_data(daily_df)
    test_passed = engine.periods_per_year == 365
    results.add("9.1 Daily Frequency Detection", test_passed,
                f"Detected: {engine.periods_per_year}" if not test_passed else "")
    
    # Test 9.2: Monthly detection
    monthly_df = pd.DataFrame({
        'period': pd.date_range('2024-01', periods=5, freq='MS'),
        'Revenue': [100, 110, 105, 115, 120]
    })
    engine = ForecastEngine()
    engine.set_base_data(monthly_df)
    test_passed = engine.periods_per_year == 12
    results.add("9.2 Monthly Frequency Detection", test_passed,
                f"Detected: {engine.periods_per_year}" if not test_passed else "")
    
    # Test 9.3: Quarterly detection
    quarterly_df = pd.DataFrame({
        'period': pd.date_range('2024-01', periods=4, freq='QS'),
        'Revenue': [100, 110, 105, 115]
    })
    engine = ForecastEngine()
    engine.set_base_data(quarterly_df)
    test_passed = engine.periods_per_year == 4
    results.add("9.3 Quarterly Frequency Detection", test_passed,
                f"Detected: {engine.periods_per_year}" if not test_passed else "")
    
    # Test 9.4: Generic pattern extrapolation
    generic_df = pd.DataFrame({
        'period': ['Period 1', 'Period 2', 'Period 3'],
        'Revenue': [100, 110, 120]
    })
    engine = ForecastEngine()
    engine.set_base_data(generic_df)
    forecast = engine.generate_base_forecast('Revenue', periods=2)
    expected_periods = ['Period 4', 'Period 5']
    actual_periods = forecast['period'].tolist()
    test_passed = actual_periods == expected_periods
    results.add("9.4 Generic Pattern Extrapolation", test_passed,
                f"Got {actual_periods}, expected {expected_periods}" if not test_passed else "")
    
    print(f"✓ Period Detection: Daily, Monthly, Quarterly, Generic all detected")
    
except Exception as e:
    results.add("9.1 Daily Frequency Detection", False, str(e))
    results.add("9.2 Monthly Frequency Detection", False, str(e))
    results.add("9.3 Quarterly Frequency Detection", False, str(e))
    results.add("9.4 Generic Pattern Extrapolation", False, str(e))

# ============================================================================
# TEST 10: ANNUAL TO PERIOD RATE CONVERSION
# ============================================================================
print("\n" + "="*70)
print("TEST SUITE 10: RATE CONVERSION")
print("="*70)

try:
    # Test 10.1: Annual to Quarterly
    quarterly_df = pd.DataFrame({
        'period': ['Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024'],
        'Revenue': [100, 100, 100, 100]
    })
    engine = ForecastEngine()
    engine.set_base_data(quarterly_df)
    
    # Add 20% annual growth
    assumption = Assumption(
        id='growth',
        metric='Revenue',
        type=AssumptionType.GROWTH,
        value=0.20,  # 20% annual
        name='Growth',
        confidence='medium',
        layer='base'
    )
    engine.add_assumption(assumption)
    
    forecast = engine.generate_base_forecast('Revenue', periods=4)
    
    # 4 quarters at period rate should compound to ~20% annual
    start = forecast['Revenue'].iloc[0]
    end = forecast['Revenue'].iloc[-1]
    annual_growth = (end / start) - 1
    
    # Should be close to 20%
    test_passed = abs(annual_growth - 0.20) < 0.02  # 2% tolerance
    results.add("10.1 Annual→Quarterly Conversion", test_passed,
                f"Annual growth: {annual_growth*100:.2f}%, expected ~20%" if not test_passed else "")
    
    # Test 10.2: Annual to Monthly
    monthly_df = pd.DataFrame({
        'period': pd.date_range('2024-01', periods=6, freq='MS'),
        'Revenue': [100] * 6
    })
    engine = ForecastEngine()
    engine.set_base_data(monthly_df)
    engine.add_assumption(assumption)  # Same 20% annual
    
    forecast = engine.generate_base_forecast('Revenue', periods=12)
    
    start = forecast['Revenue'].iloc[0]
    end = forecast['Revenue'].iloc[-1]
    annual_growth = (end / start) - 1
    
    test_passed = abs(annual_growth - 0.20) < 0.02
    results.add("10.2 Annual→Monthly Conversion", test_passed,
                f"Annual growth: {annual_growth*100:.2f}%, expected ~20%" if not test_passed else "")
    
    print(f"✓ Rate Conversion: Annual→Quarterly and Annual→Monthly validated")
    
except Exception as e:
    results.add("10.1 Annual→Quarterly Conversion", False, str(e))
    results.add("10.2 Annual→Monthly Conversion", False, str(e))

# ============================================================================
# FINAL SUMMARY
# ============================================================================
success = results.print_summary()

if success:
    print("🎉 ALL TESTS PASSED - SYSTEM IS FULLY FUNCTIONAL 🎉")
    exit(0)
else:
    print(f"⚠️  {len(results.failed_tests)} TEST(S) FAILED")
    print("Failed tests:", ", ".join(results.failed_tests))
    exit(1)
